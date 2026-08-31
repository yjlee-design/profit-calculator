# -*- coding: utf-8 -*-
"""
이익률 계산기 — Streamlit 웹앱

실행:  웹앱실행.bat  더블클릭
       또는  streamlit run app.py

계산 로직은 엔진.py 를 그대로 쓴다 (배치 버전과 결과 동일).
"""

import io
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st

BASE = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE))

import engine as E  # noqa: E402
import auth  # noqa: E402
import db  # noqa: E402

MASTER = BASE / "이익률 마스터.xlsx"
OVERRIDE = BASE / "원가보정.xlsx"
CONFIG = BASE / "입력" / "설정.xlsx"
OUT_DIR = BASE / "결과"

st.set_page_config(page_title="이익률 계산기", page_icon="📊", layout="wide")

WON = "{:,.0f}"

# ---------------------------------------------------------------- 화면 스타일
def apply_theme():
    """theme.css 를 읽어 적용 (없으면 기본 모양 그대로)"""
    f = BASE / "theme.css"
    if f.exists():
        st.markdown("<style>{}</style>".format(f.read_text(encoding="utf-8")),
                    unsafe_allow_html=True)


apply_theme()

# ---- 비밀번호를 통과해야 아래 화면이 보인다
if not auth.require_password():
    st.stop()

USE_DB = db.enabled()



# ---------------------------------------------------------------- 설정 읽기/쓰기
@st.cache_data(show_spinner=False)
def read_config_defaults(mtime):
    """입력/설정.xlsx 에서 기본값을 읽어온다 (없으면 엔진 기본값)"""
    sources = [("마진율표", "2026년 유통 마진율 및 예외단가리스트.xlsx"),
               ("구마스터", "이익률 마스터.xlsx")]
    years = E.DEFAULT_YEARS
    channels = [dict(c) for c in E.DEFAULT_CHANNELS]
    cards_all = {}          # {기준월: [(결제수단, 금액, 요율), ...]}
    samples_all = {}        # {기준월: 금액}
    period = datetime.now().strftime("%Y-%m")

    # DB 가 연결돼 있으면 DB 를 먼저 본다 (클라우드에는 설정.xlsx 가 없다)
    if db.enabled():
        try:
            ch = db.load_channels()
            if ch:
                channels = ch
            cards_all = db.load_cards() or {}
            samples_all = db.load_samples() or {}
            period = db.get_setting("period") or period
            return sources, years, channels, cards_all, samples_all, period
        except Exception:
            pass          # DB 가 잠깐 안 되면 파일로 넘어간다

    if not CONFIG.exists():
        return sources, years, channels, cards_all, samples_all, period
    wb = E.open_wb(CONFIG, data_only=True)
    try:
        if "기준파일" in wb.sheetnames:
            ws = wb["기준파일"]
            hr, col = E.find_header(ws, {"name": ["이름"], "path": ["파일경로"]})
            if hr is not None:
                for _f, _al in (("use", ["사용"]), ("years", ["최근연도수"])):
                    _, o2 = E.find_header(ws, {_f: _al})
                    if o2:
                        col.update(o2)
                found = []
                for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                    p = str(E._pick(row, col, "path") or "").strip()
                    if not p:
                        continue
                    if "use" in col and not E.is_yes(E._pick(row, col, "use")):
                        continue
                    nm = str(E._pick(row, col, "name") or "").strip() or Path(p).stem
                    found.append((nm, p))
                    if "years" in col:
                        _n = E.num(E._pick(row, col, "years"))
                        if _n >= 1:
                            years = int(_n)
                if found:
                    sources = found
        if "채널설정" in wb.sheetnames:
            ws = wb["채널설정"]
            hr, col = E.find_header(ws, {
                "name": ["채널명"], "file": ["파일명"], "seller": ["셀러수수료적용"],
                "card": ["카드수수료적용"], "sample": ["샘플비용적용"]})
            if hr is not None:
                _, opt = E.find_header(ws, {"base": ["셀러수수료기본율"]})
                if opt:
                    col.update(opt)
                found = []
                for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                    nm = E._pick(row, col, "name")
                    fn = str(E._pick(row, col, "file") or "").strip()
                    if not nm or not str(nm).strip():
                        continue
                    nm = str(nm).strip()
                    if nm.startswith("*") or not fn:
                        continue
                    found.append({
                        "name": nm, "file": fn,
                        "seller": E.is_yes(E._pick(row, col, "seller")),
                        "card": E.is_yes(E._pick(row, col, "card")),
                        "sample": E.is_yes(E._pick(row, col, "sample")),
                        "fee_base": (E.rate(E._pick(row, col, "base")) or 0.0) if "base" in col else 0.0,
                    })
                if found:
                    channels = found

        if "카드수수료" in wb.sheetnames:
            ws = wb["카드수수료"]
            hr, col = E.find_header(ws, {"method": ["결제수단"],
                                         "amount": ["정상금액", "결제금액", "금액"],
                                         "rate": ["수수료율", "수수료"]})
            if hr is not None:
                _, opt = E.find_header(ws, {"month": ["기준월"]})
                if opt:
                    col.update(opt)
                for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                    m = E._pick(row, col, "method")
                    if not m or not str(m).strip() or E.norm(m) == "합계":
                        continue
                    mon = str(E._pick(row, col, "month") or "").strip() if "month" in col else ""
                    cards_all.setdefault(mon, []).append(
                        (str(m).strip(), E.num(E._pick(row, col, "amount")),
                         E.rate(E._pick(row, col, "rate")) or 0.0))

        if "샘플비용" in wb.sheetnames:
            ws = wb["샘플비용"]
            hr, col = E.find_header(ws, {"item": ["항목"], "amount": ["금액"]})
            if hr is not None:
                _, opt = E.find_header(ws, {"month": ["기준월"]})
                if opt:
                    col.update(opt)
                for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                    it = E._pick(row, col, "item")
                    if not it or not str(it).strip() or E.norm(it) == "합계":
                        continue
                    mon = str(E._pick(row, col, "month") or "").strip() if "month" in col else ""
                    samples_all[mon] = samples_all.get(mon, 0.0) + E.num(E._pick(row, col, "amount"))

        if "기간" in wb.sheetnames:
            for row in wb["기간"].iter_rows(values_only=True):
                if row and E.norm(row[0]) == "기준월":
                    for v in row[1:]:
                        if v is not None and str(v).strip():
                            period = str(v).strip()
                            break
    finally:
        wb.close()
    return sources, years, channels, cards_all, samples_all, period


def save_config(period, sources, years, ch_df, cards_now, smp_year, cards_all, samples_all):
    """현재 화면 값을 입력/설정.xlsx 로 저장.
       카드수수료·샘플비용은 **다른 달 값을 지우지 않고** 이번 달분만 갱신한다."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    HDR = PatternFill("solid", fgColor="D9E1F2")
    IN = PatternFill("solid", fgColor="FFF2CC")

    def head(ws, row, names, widths):
        from openpyxl.utils import get_column_letter
        for i, n in enumerate(names, start=1):
            c = ws.cell(row, i, n)
            c.font, c.fill = Font(bold=True), HDR
            c.alignment = Alignment(horizontal="center")
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "기간"
    head(ws, 3, ["항목", "값"], [16, 20])
    ws.cell(4, 1, "기준월")
    ws.cell(4, 2, period).fill = IN

    ws = wb.create_sheet("기준파일")
    ws.cell(1, 1, "원가·셀러수수료를 읽어올 파일. 위에 있는 파일이 우선합니다.").font = Font(bold=True)
    ws.cell(2, 1, "* 파일명만 적으면 이 폴더 기준, 전체 경로(공유폴더·OneDrive)도 됩니다.").font = Font(size=9, color="808080")
    ws.cell(5, 1, "* 실행할 때마다 그 경로의 최신 파일을 그대로 읽습니다.").font = Font(size=9, color="808080")
    head(ws, 7, ["순서", "이름", "파일경로", "사용", "최근연도수"], [6, 16, 68, 8, 12])
    for i, (nm, p) in enumerate(sources, start=8):
        ws.cell(i, 1, i - 7)
        ws.cell(i, 2, nm)
        ws.cell(i, 3, p).fill = IN
        c = ws.cell(i, 4, "O")
        c.alignment = Alignment(horizontal="center")
        c.fill = IN
        if i == 8:
            yc = ws.cell(i, 5, int(years))
            yc.alignment = Alignment(horizontal="center")
            yc.fill = IN

    ws = wb.create_sheet("채널설정")
    head(ws, 3, ["채널명", "파일명", "셀러수수료적용", "카드수수료적용", "샘플비용적용",
                 "셀러수수료기본율"], [16, 22, 15, 15, 14, 16])
    for i, (_, row) in enumerate(ch_df.iterrows(), start=4):
        ws.cell(i, 1, str(row["채널명"]))
        ws.cell(i, 2, str(row["파일명"]))
        for j, k in enumerate(["셀러수수료", "카드수수료", "샘플비용"], start=3):
            c = ws.cell(i, j, "O" if bool(row[k]) else "X")
            c.alignment = Alignment(horizontal="center")
            c.fill = IN
        c = ws.cell(i, 6, float(row["셀러수수료 기본율"]))
        c.number_format, c.fill = "0.00%", IN

    # 카드수수료 — 기존 월은 그대로 두고 이번 달만 교체
    merged_cards = {k: list(v) for k, v in (cards_all or {}).items()}
    merged_cards[period] = [(m, a, r) for m, a, r in cards_now]
    ws = wb.create_sheet("카드수수료")
    ws.cell(1, 1, "스룩 결제수단별 금액 (월별로 쌓입니다)").font = Font(bold=True)
    ws.cell(2, 1, "* 웹앱에서 입력하면 자동으로 여기에 저장됩니다.").font = Font(size=9, color="808080")
    head(ws, 4, ["기준월", "결제수단", "정상금액", "수수료율"], [12, 16, 16, 12])
    i = 5
    for mon in sorted(merged_cards, key=lambda x: (x == "", x)):
        for m, a, r in merged_cards[mon]:
            ws.cell(i, 1, mon)
            ws.cell(i, 2, m)
            c = ws.cell(i, 3, float(a))
            c.number_format, c.fill = "#,##0", IN
            c = ws.cell(i, 4, float(r))
            c.number_format, c.fill = "0.00%", IN
            i += 1
    ws.freeze_panes = "A5"

    # 샘플비용 — 화면의 12개월 값으로 그 해를 갱신, 다른 해는 유지
    merged_smp = dict(samples_all or {})
    for mon, amt in (smp_year or {}).items():
        merged_smp[mon] = amt
    ws = wb.create_sheet("샘플비용")
    ws.cell(1, 1, "자체 부담 비용 (월별로 쌓입니다)").font = Font(bold=True)
    ws.cell(2, 1, "* 웹앱에서 입력하면 자동으로 여기에 저장됩니다. 지난 달 값은 지우지 마세요.").font = Font(size=9, color="808080")
    head(ws, 4, ["기준월", "항목", "금액"], [12, 26, 16])
    i = 5
    for mon in sorted(merged_smp, key=lambda x: (x == "", x)):
        ws.cell(i, 1, mon)
        ws.cell(i, 2, "자체샘플+이벤트지원")
        c = ws.cell(i, 3, float(merged_smp[mon]))
        c.number_format, c.fill = "#,##0", IN
        i += 1
    ws.freeze_panes = "A5"

    CONFIG.parent.mkdir(exist_ok=True)
    wb.save(CONFIG)


def append_overrides(typed, names):
    """화면에서 넣은 원가를 원가보정.xlsx [원가보정] 시트에 추가/갱신한다.
       우선적용은 X (기준 파일에 값이 없을 때만 쓰임)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    IN = PatternFill("solid", fgColor="FFF2CC")
    today = datetime.now().strftime("%y.%m.%d")

    if OVERRIDE.exists():
        wb = openpyxl.load_workbook(OVERRIDE)
        ws = wb["원가보정"] if "원가보정" in wb.sheetnames else wb.worksheets[0]
        hr, col = E.find_header(ws, {"code": ["상품코드"], "cost": ["원가"]})
        for f, al in (("force", ["우선적용"]), ("name", ["품목명(참고)", "품목명"]),
                      ("note", ["비고"])):
            _, o = E.find_header(ws, {f: al})
            if o:
                col.update(o)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "원가보정"
        ws.cell(1, 1, "기준 파일에 없는 원가를 여기에 적습니다").font = Font(bold=True)
        for i, h in enumerate(["상품코드", "원가", "우선적용", "품목명(참고)", "채널", "비고"], 1):
            ws.cell(5, i, h).font = Font(bold=True)
        hr, col = 5, {"code": 1, "cost": 2, "force": 3, "name": 4, "note": 6}

    # 이미 있는 코드는 값만 갱신
    where = {}
    for r in range(hr + 1, ws.max_row + 1):
        k = E.code_key(ws.cell(r, col["code"]).value)
        if k:
            where[k] = r
    nxt = ws.max_row + 1
    n = 0
    for k, v in typed.items():
        raw, nm = names.get(k, (k, ""))
        r = where.get(k)
        if r is None:
            r, nxt = nxt, nxt + 1
            ws.cell(r, col["code"], raw)
        c = ws.cell(r, col["cost"], float(v))
        c.number_format, c.fill = "#,##0", IN
        if "force" in col:
            fc = ws.cell(r, col["force"], "X")
            fc.alignment = Alignment(horizontal="center")
            fc.fill = IN
        if "name" in col and nm:
            ws.cell(r, col["name"], nm)
        if "note" in col:
            ws.cell(r, col["note"], "{} 화면에서 직접 입력".format(today))
        n += 1
    wb.save(OVERRIDE)
    return n


def save_report(period, xlsx, summary=""):
    """그 달의 결과 엑셀을 보관"""
    fname = "{} 이익률.xlsx".format(period)
    if db.enabled():
        try:
            db.save_month_report(period, fname, xlsx, summary)
            return
        except Exception:
            pass
    try:
        OUT_DIR.mkdir(exist_ok=True)
        (OUT_DIR / fname).write_bytes(xlsx)
    except OSError:
        pass


def load_report(period):
    """(파일명, bytes, 요약, 저장시각) 또는 None"""
    if db.enabled():
        try:
            got = db.load_month_report(period)
            if got:
                return got
        except Exception:
            pass
    f = OUT_DIR / "{} 이익률.xlsx".format(period)
    if f.exists():
        import datetime as _dt
        return (f.name, f.read_bytes(), "",
                _dt.datetime.fromtimestamp(f.stat().st_mtime))
    return None


def list_reports():
    if db.enabled():
        try:
            return db.list_month_reports()
        except Exception:
            pass
    import datetime as _dt, re as _re
    out = []
    if OUT_DIR.is_dir():
        for f in sorted(OUT_DIR.glob("* 이익률.xlsx"), reverse=True):
            m = _re.match(r"(\d{4}-\d{2})", f.name)
            if m:
                out.append((m.group(1), f.name, "",
                            _dt.datetime.fromtimestamp(f.stat().st_mtime)))
    return out


INPUT_DIR_SAVE = BASE / "결과" / "입력보관"


def save_month_files(period, files):
    """그 달에 쓴 매출 파일을 보관 — 페이지를 나갔다 와도 결과가 남도록"""
    if not files:
        return
    if db.enabled():
        try:
            db.save_month_files(period, files)
            return
        except Exception:
            pass
    try:
        d = INPUT_DIR_SAVE / period
        d.mkdir(parents=True, exist_ok=True)
        for old in d.glob("*"):
            old.unlink()
        for ch, (fn, data) in files.items():
            (d / "{}__{}".format(ch, fn)).write_bytes(data)
    except OSError:
        pass


def load_month_files(period):
    """{채널명: (파일명, bytes)}  없으면 {}"""
    if db.enabled():
        try:
            got = db.load_month_files(period)
            if got:
                return got
        except Exception:
            pass
    out = {}
    try:
        d = INPUT_DIR_SAVE / period
        if d.is_dir():
            for f in d.glob("*__*"):
                ch, fn = f.name.split("__", 1)
                out[ch] = (fn, f.read_bytes())
    except OSError:
        pass
    return out


HISTORY_FILE = BASE / "결과" / "월별추이.json"


def save_history(period, results, total):
    """이 달 결과를 이력에 남긴다 (DB 우선, 없으면 결과/월별추이.json)"""
    rows = [{"period": period, "channel": r["ch"]["name"], "sales": r["sales"],
             "cost": r["cost"], "fee": r["fee"], "card": r.get("card", 0),
             "sample": r.get("sample", 0), "profit": r["profit"],
             "delivery": r["delivery_sales"]} for r in results]
    rows.append({"period": period, "channel": "전체", "sales": total["sales"],
                 "cost": total["cost"], "fee": total["fee"], "card": total["card"],
                 "sample": total["sample"], "profit": total["profit"],
                 "delivery": total["delivery"]})
    if db.enabled():
        try:
            db.save_month_result(period, results, total)
            return
        except Exception:
            pass
    try:
        import json
        old = []
        if HISTORY_FILE.exists():
            old = [x for x in json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
                   if x.get("period") != period]
        HISTORY_FILE.parent.mkdir(exist_ok=True)
        HISTORY_FILE.write_text(json.dumps(old + rows, ensure_ascii=False, indent=1),
                                encoding="utf-8")
    except OSError:
        pass


def load_history():
    if db.enabled():
        try:
            return db.load_month_results()
        except Exception:
            pass
    try:
        import json
        if HISTORY_FILE.exists():
            return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        pass
    return []


# ---------------------------------------------------------------- 기준 파일 로드 (캐시)
@st.cache_data(show_spinner="기준 파일 읽는 중...")
def load_lookups_cached(blobs):
    """blobs: ((라벨, 바이트, 태그), ...)  — 태그가 바뀌면 다시 읽는다"""
    return E.load_lookups([(lb, io.BytesIO(b)) for lb, b, _t in blobs])


@st.cache_data(show_spinner=False)
def load_override_cached(data, _tag):
    if data is None:
        return {}, {}
    return E.load_overrides(io.BytesIO(data))


def resolve_path(p):
    """설정에 적힌 경로 → 실제 경로 (상대경로는 이 폴더 기준)"""
    f = Path(str(p).strip().strip('"'))
    return f if f.is_absolute() else (BASE / f)


def stat_source(p):
    """경로를 '들여다보기만' 한다 (읽지는 않음). 구글드라이브·공유폴더에서도 빠르다."""
    try:
        f = resolve_path(p)
        if not f.exists():
            return {"path": f, "ok": False, "why": "파일 없음", "mtime": 0, "size": 0}
        s = f.stat()
        return {"path": f, "ok": True, "why": "", "mtime": s.st_mtime, "size": s.st_size}
    except OSError as e:
        return {"path": Path(str(p)), "ok": False, "why": "읽기 실패 ({})".format(e),
                "mtime": 0, "size": 0}


def when(ts):
    return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M") if ts else "-"


# ---------------------------------------------------------------- 사이드바
NAV = {"데이터 입력": "input", "결과": "result"}
NAV_REV = {v: k for k, v in NAV.items()}


def go(view):
    """메뉴와 화면을 함께 바꾼다.
       라디오는 이미 그려진 뒤라 직접 못 고치므로, 다음 실행에 반영되도록 예약한다.
       호출한 쪽에서 st.rerun() 을 해야 한다."""
    st.session_state["view"] = view
    st.session_state["_nav_pending"] = NAV_REV[view]


st.sidebar.title("📊 이익률 계산기")

defaults_key = CONFIG.stat().st_mtime if CONFIG.exists() else 0
def_sources, def_years, def_channels, def_cards_all, def_samples_all, def_period = read_config_defaults(defaults_key)

def split_period(p, fallback=None):
    """'2026-07' → (2026, 7). 이상하면 fallback(또는 오늘)"""
    try:
        return int(str(p)[:4]), int(str(p)[5:7])
    except (ValueError, TypeError):
        n = fallback or datetime.now()
        return n.year, n.month


if "period_year" not in st.session_state:
    _y, _m = split_period(def_period)
    st.session_state["period_year"] = _y
    st.session_state["period_month"] = _m
if "period_override" in st.session_state:            # 월 버튼으로 옮겨온 경우
    _y, _m = split_period(st.session_state.pop("period_override"))
    st.session_state["period_year"] = _y
    st.session_state["period_month"] = _m

_now = datetime.now().year
_years = {int(k[:4]) for k in list(def_cards_all) + list(def_samples_all)
          if len(str(k)) >= 4 and str(k)[:4].isdigit()}
_years |= {_now - 1, _now, _now + 1, st.session_state["period_year"]}
_years = sorted(_years)

st.sidebar.markdown("**기준월**")
_c1, _c2 = st.sidebar.columns(2)
_yy = _c1.selectbox("년", _years, key="period_year",
                    format_func=lambda x: "{}년".format(x), label_visibility="collapsed")
_mm = _c2.selectbox("월", list(range(1, 13)), key="period_month",
                    format_func=lambda x: "{}월".format(x), label_visibility="collapsed")
period = "{:04d}-{:02d}".format(_yy, _mm)
st.sidebar.caption("결과 파일 이름이 됩니다 → `{} 이익률.xlsx`".format(period))

st.sidebar.divider()

# ---- 왼쪽 상단 메뉴
if "_nav_pending" in st.session_state:
    st.session_state["nav_menu"] = st.session_state.pop("_nav_pending")
if "nav_menu" not in st.session_state:
    st.session_state["nav_menu"] = NAV_REV.get(st.session_state.get("view", "input"),
                                               "데이터 입력")
_pick = st.sidebar.radio("메뉴", list(NAV), key="nav_menu", label_visibility="collapsed")
st.session_state["view"] = NAV[_pick]
if NAV[_pick] == "result" and not st.session_state.get("out"):
    st.sidebar.caption("아직 계산 결과가 없습니다. **데이터 입력** 에서 계산하세요.")
st.sidebar.divider()


# 기준월에 해당하는 값 (없으면 가장 최근 달의 결제수단 구성을 금액 0으로 가져옴)
_extra = st.session_state.get("extra_cards", [])
if period in def_cards_all:
    def_cards = list(def_cards_all[period])
elif def_cards_all:
    _last = sorted(def_cards_all)[-1]
    def_cards = [(m, 0.0, r) for m, _a, r in def_cards_all[_last]]
else:
    def_cards = [(m, 0.0, r) for m, r in E.DEFAULT_CARDS]
_have = {m for m, _a, _r in def_cards}
for m in _extra:
    if m not in _have:
        def_cards.append((m, 0.0, 0.0275))
def_samples_by_month = dict(def_samples_all)

# ---------------------------------------------------------------- 기준 파일
# UI 는 '데이터 입력' 화면 위쪽에 그리고, 조회표 적재는 어느 화면에서든 돌아야 한다.
if USE_DB:
    _expanded, _src_notes, stats = [], [], []
else:
    _expanded, _src_notes = E.expand_sources(def_sources, BASE, def_years)
    stats = [(nm, str(f), stat_source(f)) for nm, f in _expanded]


def source_sig(uploaded=None):
    return ((period,)
            + tuple((nm, str(s["path"]), s["mtime"], s["size"]) for nm, _p, s in stats)
            + tuple((f.name, f.size) for f in (uploaded or [])))


def load_sources(uploaded=None, box=None):
    """기준 원가·요율을 읽어 세션에 담는다. box 가 있으면 오류를 거기 표시."""
    err = (box or st).error
    blobs = [("업로드:" + f.name, f.getvalue()) for f in (uploaded or [])]
    for nm, _p, s in stats:
        if s["ok"]:
            try:
                blobs.append((nm, s["path"].read_bytes()))
            except OSError as e:
                err("{} — 읽기 실패 ({})".format(nm, e))
    try:
        if USE_DB and not blobs:
            with st.spinner("기준 원가 불러오는 중..."):
                d_cost, d_fee, d_origin = db.load_lookups_asof(period)
            st.session_state["lookups"] = {
                "sig": source_sig(uploaded), "cost": d_cost, "fee": d_fee, "conflicts": [],
                "report": [{"label": "Supabase", "format": "DB", "cost": len(d_cost),
                            "fee": len(d_fee), "cost_new": len(d_cost), "fee_new": len(d_fee)}],
                "origin": d_origin, "at": datetime.now()}
        elif blobs:
            with st.spinner("기준 파일 읽는 중..."):
                c, f, cf, rep, org = E.load_lookups(
                    [(lb, io.BytesIO(b)) for lb, b in blobs], E.month_cutoff(period))
            st.session_state["lookups"] = {
                "sig": source_sig(uploaded), "cost": c, "fee": f, "conflicts": cf,
                "report": rep, "origin": org, "at": datetime.now()}
    except Exception as e:
        err(str(e))


# 아직 한 번도 안 읽었으면 어느 화면이든 먼저 읽어둔다
if st.session_state.get("lookups") is None:
    load_sources()

loaded = st.session_state.get("lookups")
cost = fee = conflicts = src_report = origin = None
if loaded:
    cost, fee = loaded["cost"], loaded["fee"]
    conflicts, src_report, origin = loaded["conflicts"], loaded["report"], loaded["origin"]


def render_source_panel():
    """데이터 입력 화면 위쪽에 그리는 기준 파일 패널"""
    global cost, fee, conflicts, src_report, origin, loaded
    with st.expander("📁 기준 파일 — 원가·셀러수수료를 어디서 읽는지", expanded=False):
        c1, c2 = st.columns([3, 2])
        with c1:
            up_ = st.file_uploader("기준 파일 직접 올리기 (선택 · 최우선 적용)",
                                   type=["xlsx", "xlsm"], accept_multiple_files=True,
                                   key="up_src")
            for nm, p, s in stats:
                if s["ok"]:
                    st.caption("**{}** · 파일 수정 {}".format(nm, when(s["mtime"])))
                else:
                    st.error("{} — {}  ({})".format(nm, s["why"], p))
            for _n in _src_notes:
                if "없" in _n or "못" in _n:
                    st.error(_n)
        with c2:
            b1, b2 = st.columns(2)
            do_up = b1.button("🔄 업데이트", width="stretch",
                              help="기준 파일을 지금 다시 읽습니다")
            if b2.button("상태 새로고침", width="stretch",
                         help="파일이 바뀌었는지만 확인"):
                st.rerun()
            cur = st.session_state.get("lookups")
            if cur is not None and cur["sig"] != source_sig(up_) and not do_up:
                st.warning("기준 파일이 바뀌었습니다. **🔄 업데이트** 를 눌러 반영하세요.")
            if do_up:
                load_sources(up_)
                st.rerun()
            if src_report:
                for r in src_report:
                    st.caption("· {} — 원가 {:,} (신규 {:,}) · 요율 {:,}".format(
                        r["label"], r["cost"], r["cost_new"], r["fee"]))
                st.success("합계 원가 {:,}건 · 셀러수수료 {:,}건".format(len(cost), len(fee)))
                st.caption("읽은 시각 {}".format(cur["at"].strftime("%m-%d %H:%M:%S")))
            elif not stats and not USE_DB:
                st.error("읽을 기준 파일이 없습니다. 입력/설정.xlsx [기준파일] 시트를 확인하세요.")


st.sidebar.divider()
ov_bytes, ov_tag = None, ""
if USE_DB:
    try:
        override, fee_override = db.load_overrides()
    except Exception as e:
        override, fee_override = {}, {}
        st.sidebar.error("보정 읽기 실패: {}".format(e))
    st.sidebar.caption("보정(DB) — 원가 {:,}건 · 요율 {:,}건".format(len(override), len(fee_override)))
elif OVERRIDE.exists() and st.sidebar.toggle("원가보정.xlsx 사용", value=True):
    ov_bytes = OVERRIDE.read_bytes()
    ov_tag = "{}:{}".format(OVERRIDE, OVERRIDE.stat().st_mtime)
    override, fee_override = load_override_cached(ov_bytes, ov_tag)
    st.sidebar.caption("보정 — 원가 {:,}건 · 요율 {:,}건".format(len(override), len(fee_override)))
else:
    override, fee_override = load_override_cached(None, "")


st.sidebar.divider()
if USE_DB:
    _ls = None
    try:
        _ls = db.last_sync()
    except Exception:
        pass
    st.sidebar.success("☁️ Supabase 연결됨")
    if _ls:
        st.sidebar.caption("마지막 동기화 {} · {}".format(
            _ls["synced_at"].strftime("%Y-%m-%d %H:%M"), _ls["detail"]))
    else:
        st.sidebar.caption("아직 동기화 기록이 없습니다. 사무실 PC 에서 `동기화.bat` 을 실행하세요.")
else:
    _prob = None
    try:
        _prob = db.url_problem()
    except Exception:
        pass
    if _prob:
        st.sidebar.warning("DB 미연결 — " + _prob)
    st.sidebar.caption("💾 파일 모드 (이 PC 의 엑셀 파일 사용)")
auth.logout_button()


# ---------------------------------------------------------------- 본문
def run_calc(snap=None):
    """저장된 입력값으로 계산. 화면에서 직접 넣은 원가(manual_cost)도 함께 적용.
       반환: (성공여부, 사유목록)"""
    snap = snap or st.session_state.get("calc_snapshot")
    if not snap:
        return False, ["계산에 쓸 입력값이 없습니다. 입력 화면에서 다시 계산해 주세요."]

    manual = st.session_state.get("manual_cost", {})
    ov = dict(snap["override"])
    for k, v in manual.items():
        if k and v:
            ov.setdefault(k, (float(v), False))     # 기준 파일에 없을 때만 쓰임

    results, errs = [], []
    for ch in snap["channels"]:
        got = snap["files"].get(ch["name"])
        if got is None:
            errs.append("{} — 파일 없음 (건너뜀)".format(ch["name"]))
            continue
        fname, data = got
        try:
            rows = E.read_sales(io.BytesIO(data), fname)
        except Exception as e:
            errs.append("{} ({}) — {}".format(ch["name"], fname, e))
            continue
        results.append(E.calc_channel(ch, rows, snap["cost"], snap["fee"],
                                      ov, snap["fee_override"], snap["origin"]))
    if not results:
        st.session_state.pop("out", None)
        return False, errs

    total = E.apply_totals(results, snap["card_total"], snap["sample"])
    wb = E.build_report(results, total, snap["cards"], snap["card_total"],
                        snap["sample_items"], snap["sample"], snap["conflicts"],
                        snap["period"])
    buf = io.BytesIO()
    wb.save(buf)
    st.session_state["out"] = {
        "results": results, "total": total, "errs": errs,
        "xlsx": buf.getvalue(), "period": snap["period"],
        "manual_used": {k: v for k, v in manual.items() if v},
    }
    save_history(snap["period"], results, total)     # 월별 추이 그래프용 이력
    save_month_files(snap["period"], snap["files"])   # 다음에 열어도 결과가 보이도록
    save_report(snap["period"], buf.getvalue(),       # 결과 엑셀도 달마다 보관
                "매출 {:,.0f} / 이익 {:,.0f} / {:.2f}%".format(
                    total["gross"], total["profit"], total["margin"] * 100))
    return True, errs


_prev = (st.session_state.get("calc_snapshot") or {}).get("files", {})

# ---- 페이지를 새로 열었을 때, 그 달의 저장된 결과를 되살린다
if not st.session_state.get("out") and not st.session_state.get("restore_done"):
    st.session_state["restore_done"] = True
    _saved = load_month_files(period)
    if _saved and cost:
        _snap = {
            "channels": [dict(c) for c in def_channels], "files": _saved,
            "cost": cost, "fee": fee, "origin": origin, "conflicts": conflicts or [],
            "override": override, "fee_override": fee_override,
            "cards": [], "card_total": 0.0,
            "sample_items": [], "sample": 0.0, "period": period,
        }
        # 그 달의 비용은 설정에서 다시 읽어 채운다
        _c = def_cards_all.get(period, [])
        _snap["cards"], _snap["card_total"] = E.card_rows_total(_c)
        _sm = float(def_samples_by_month.get(period, 0.0))
        _snap["sample_items"], _snap["sample"] = [("자체샘플+이벤트지원", _sm)], _sm
        with st.spinner("{} 저장된 결과 불러오는 중...".format(period)):
            _ok, _ = run_calc(_snap)
        if _ok:
            st.session_state["calc_snapshot"] = _snap
            st.session_state["restored_from_save"] = True
            go("result")
            st.rerun()

VIEW = st.session_state.setdefault("view", "input")

if VIEW == "result" and not st.session_state.get("out"):
    st.session_state["view"] = VIEW = "input"       # 결과가 없으면 입력 화면

if VIEW == "result":
    _o = st.session_state["out"]
    _c1, _c2 = st.columns([3, 1])
    _c1.title("{} 이익률 결과".format(_o["period"]))
    _c2.write("")
    if _c2.button("← 데이터 입력으로", width="stretch"):
        go("input")
        st.rerun()
    if st.session_state.pop("restored_from_save", False):
        st.info("저장해 둔 **{}** 자료로 다시 계산했습니다. "
                "새 자료로 바꾸려면 왼쪽 메뉴의 **데이터 입력** 으로 가서 파일을 다시 올리세요."
                .format(_o["period"]))
else:
    st.title("이익률 계산")
if VIEW == "input":
    st.caption("이카운트 매출 파일을 올리면 채널별 이익률을 계산합니다. "
               "배송비는 매출=원가로 처리하여 이익에서 제외됩니다.")

    render_source_panel()

    # ---- 채널 설정 (업로더보다 먼저 정의해야 하므로 위에 두되, 접어 둔다)
    with st.expander("⚙️ 채널 설정  —  채널을 추가하거나 비용 차감 대상을 바꿀 때만"):
        st.markdown("채널을 추가하거나 비용 차감 대상을 바꿀 때만 수정하세요. "
                    "**카드수수료·샘플비용은 총액이므로 한 채널에만 체크**하는 것이 맞습니다.")
        ch_df = st.data_editor(
            pd.DataFrame([{
                "채널명": c["name"], "파일명": c["file"],
                "셀러수수료": c["seller"], "카드수수료": c["card"], "샘플비용": c["sample"],
                "셀러수수료 기본율": c.get("fee_base", 0.0),
            } for c in def_channels]),
            num_rows="dynamic", width="stretch", key="ch_editor",
            column_config={
                "채널명": st.column_config.TextColumn(required=True),
                "파일명": st.column_config.TextColumn(help="배치 실행 시 입력/ 폴더에서 찾을 이름"),
                "셀러수수료": st.column_config.CheckboxColumn(help="상품별 셀러수수료를 차감"),
                "카드수수료": st.column_config.CheckboxColumn(help="카드수수료 총액을 차감"),
                "샘플비용": st.column_config.CheckboxColumn(help="샘플비용 총액을 차감"),
                "셀러수수료 기본율": st.column_config.NumberColumn(
                    format="percent", min_value=0.0, max_value=1.0, step=0.01,
                    help="기준 파일에 요율이 없을 때 적용 (0.15 = 15%)"),
            })
        channels = [{
            "name": str(r["채널명"]).strip(), "file": str(r["파일명"] or "").strip(),
            "seller": bool(r["셀러수수료"]), "card": bool(r["카드수수료"]),
            "sample": bool(r["샘플비용"]), "fee_base": float(r["셀러수수료 기본율"] or 0.0),
        } for _, r in ch_df.iterrows() if str(r["채널명"]).strip()]

    # ---- ① 매출 파일
    st.subheader("① 이카운트 매출 파일")
    st.caption("채널별로 다운로드한 파일을 올려주세요. 필요한 열: "
               "`품목명[규격]코드` · `수량` · `판매액`  "
               "(`품목그룹3코드` 가 있으면 배송비를 자동 분리합니다)")
    uploads = {}
    cols = st.columns(min(len(channels), 4) or 1)
    for i, ch in enumerate(channels):
        with cols[i % len(cols)]:
            uploads[ch["name"]] = st.file_uploader(
                ch["name"], type=["xlsx", "xlsm", "csv"], key="up_" + ch["name"])

    st.divider()

    # ---- ② 비용 입력 (같은 페이지)
    st.subheader("② 비용 입력")
    if True:
        year = period[:4] if len(period) >= 4 and period[:4].isdigit() else str(datetime.now().year)

        st.subheader("카드수수료  ·  {}".format(period))
        st.caption("월을 눌러 옮겨 다니며 입력하세요. 선택한 달이 **기준월**이 되어 계산에 쓰입니다. "
                   "결제수단별 **정상금액**만 채우면 수수료가 자동 계산됩니다.")

        # ---- 스룩페이 매출통계 파일로 자동 입력
        srk = st.file_uploader(
            "스룩페이 매출통계 파일 올리기  (매출/정산 → 엑셀 내려받기)",
            type=["xlsx", "xlsm"], key="up_srook_%s" % period,
            help="결제수단별 [정상금액]을 읽어 아래 칸을 자동으로 채웁니다. "
                 "수수료는 아래 요율로 계산합니다.")
        if srk is not None:
            sig = "{}:{}:{}".format(period, srk.name, srk.size)
            try:
                sk = E.read_srookpay(io.BytesIO(srk.getvalue()))
                st.session_state["srook_" + period] = sk
                if st.session_state.get("srook_applied") != sig:
                    # 위치가 아니라 **결제수단 이름**으로 맞춘다 (순서가 다를 수 있음)
                    slot = {E.norm(m): i for i, (m, _a, _r) in enumerate(def_cards)}
                    unmatched = []
                    for m, a, _f, _r in sk["rows"]:
                        i = slot.get(E.norm(m))
                        if i is None:
                            unmatched.append(m)
                            continue
                        st.session_state["card_amt_%s_%d" % (period, i)] = float(a)
                    # 파일에는 없는 결제수단은 0 으로 (지난달 값이 남지 않도록)
                    in_file = {E.norm(m) for m, _a, _f, _r in sk["rows"]}
                    for i, (m, _a, _r) in enumerate(def_cards):
                        if E.norm(m) not in in_file:
                            st.session_state["card_amt_%s_%d" % (period, i)] = 0.0
                    st.session_state["srook_unmatched"] = unmatched
                    st.session_state["srook_applied"] = sig
                    st.rerun()
            except Exception as e:
                st.error("스룩페이 파일을 읽지 못했습니다 — {}".format(e))

        sk = st.session_state.get("srook_" + period)
        if sk:
            pf, pt = sk["period_from"], sk["period_to"]
            want = period                                   # 예: 2026-08
            ok_month = (not pf) or (pf[:7] == want and pt[:7] == want)
            msg = "스룩페이 자료 **{} ~ {}** · 정상금액 합계 {} 원".format(
                pf or "?", pt or "?", WON.format(sk["total_amount"]))
            (st.success if ok_month else st.warning)(msg)
            if not ok_month:
                st.warning("이 파일의 기간이 기준월({})과 다릅니다. 월을 확인하세요.".format(period))
            _um = st.session_state.get("srook_unmatched") or []
            if _um:
                st.warning("아래 목록에 없는 결제수단이라 채우지 못했습니다: **{}**  "
                           "→ [결제수단 추가·삭제] 에서 같은 이름으로 추가하세요.".format(", ".join(_um)))

        # 월 이동 버튼 — 누르면 그 달로 기준월이 바뀐다
        mrow = st.columns(12)
        for m in range(1, 13):
            key_m = "{}-{:02d}".format(year, m)
            label = "{}월{}".format(m, " ●" if key_m == period else "")
            if mrow[m - 1].button(label, key="mtab_%s" % key_m,
                                  type="primary" if key_m == period else "secondary",
                                  width="stretch"):
                st.session_state["period_override"] = key_m
                st.rerun()
        st.write("")

        CARD_W = [2, 2, 1.1, 0.35, 2]        # 결제수단 / 정상금액 / 수수료율 / % / 수수료
        h = st.columns(CARD_W)
        for c, t in zip(h, ["결제수단", "정상금액", "수수료율", "", "수수료"]):
            c.markdown("**{}**".format(t) if t else "")

        cards_in, amt_sum, fee_sum = [], 0.0, 0.0
        for i, (m, a, r) in enumerate(def_cards):
            c = st.columns(CARD_W)
            c[0].markdown("<div style='padding-top:.55rem'>{}</div>".format(m), unsafe_allow_html=True)
            amt = c[1].number_input("정상금액 " + m, min_value=0.0, value=float(a), step=1000.0,
                                    format="%.0f", key="card_amt_%s_%d" % (period, i),
                                    label_visibility="collapsed")
            # 수수료율은 퍼센트로 입력받는다 (2.75 → 0.0275)
            pct = c[2].number_input("수수료율 " + m, min_value=0.0, max_value=100.0,
                                    value=round(float(r) * 100, 4), step=0.05, format="%.2f",
                                    key="card_rt_%s_%d" % (period, i), label_visibility="collapsed")
            c[3].markdown("<div style='padding-top:.55rem;color:#64748B'>%</div>",
                          unsafe_allow_html=True)
            rt = pct / 100.0
            f = amt * rt
            c[4].markdown("<div style='padding-top:.55rem;text-align:right'>{}</div>".format(
                WON.format(f) if f else "-"), unsafe_allow_html=True)
            cards_in.append((m, amt, rt))
            amt_sum += amt
            fee_sum += f

        t = st.columns(CARD_W)
        t[0].markdown("**합계**")
        t[1].markdown("<div style='text-align:right'><b>{}</b></div>".format(WON.format(amt_sum)),
                      unsafe_allow_html=True)
        t[4].markdown("<div style='text-align:right'><b>{}</b></div>".format(WON.format(fee_sum)),
                      unsafe_allow_html=True)
        cards, card_total = E.card_rows_total(cards_in)


        with st.expander("결제수단 추가·삭제"):
            new_m = st.text_input("추가할 결제수단 이름", key="card_new")
            cc = st.columns(2)
            if cc[0].button("추가", key="card_add") and new_m.strip():
                st.session_state["extra_cards"] = st.session_state.get("extra_cards", []) + [new_m.strip()]
                st.rerun()
            if cc[1].button("추가한 결제수단 모두 지우기", key="card_clr"):
                st.session_state["extra_cards"] = []
                st.rerun()
            st.caption("추가한 결제수단은 저장하면 설정.xlsx 에 남습니다.")

        st.divider()

        st.subheader("샘플비용  ·  {}".format(period))
        st.caption("월을 눌러 옮겨 다니며 입력하세요. **기준월({})** 의 금액이 이번 계산에 차감됩니다. "
                   "지난 달 값은 그대로 쌓입니다.".format(period))

        # 월 이동 버튼 (카드수수료와 동일 — 기준월이 함께 바뀝니다)
        srow = st.columns(12)
        for mth in range(1, 13):
            key_m = "{}-{:02d}".format(year, mth)
            if srow[mth - 1].button("{}월{}".format(mth, " ●" if key_m == period else ""),
                                    key="smtab_%s" % key_m,
                                    type="primary" if key_m == period else "secondary",
                                    width="stretch"):
                st.session_state["period_override"] = key_m
                st.rerun()
        st.write("")

        # 선택한 달만 입력
        smp_year = {"{}-{:02d}".format(year, m): float(def_samples_by_month.get(
            "{}-{:02d}".format(year, m), 0.0)) for m in range(1, 13)}
        sc = st.columns([2, 2, 3])
        sc[0].markdown("<div style='padding-top:.55rem'><b>자체샘플+이벤트지원</b></div>",
                       unsafe_allow_html=True)
        v_now = sc[1].number_input("샘플비용 " + period, min_value=0.0,
                                   value=float(def_samples_by_month.get(period, 0.0)),
                                   step=1000.0, format="%.0f", key="smp_%s" % period,
                                   label_visibility="collapsed")
        smp_year[period] = v_now
        cum = sum(smp_year.values())
        sc[2].markdown("<div style='padding-top:.55rem;color:#64748B'>"
                       "{}년 누적 <b style='color:#0F172A'>{}</b> 원</div>".format(
                           year[2:], WON.format(cum)), unsafe_allow_html=True)

        with st.expander("{}년 12개월 한눈에 보기".format(year[2:])):
            st.dataframe(
                pd.DataFrame([{"월": "{}월".format(m),
                               "자체샘플+이벤트지원": smp_year["{}-{:02d}".format(year, m)],
                               "": "◀ 이번 계산" if "{}-{:02d}".format(year, m) == period else ""}
                              for m in range(1, 13)]),
                width="stretch", hide_index=True,
                column_config={"자체샘플+이벤트지원": st.column_config.NumberColumn(format="%d")})
            st.caption("누적 {} 원".format(WON.format(cum)))

        sample = smp_year.get(period, 0.0)
        sample_items = [("자체샘플+이벤트지원", sample)]
        st.metric("{} 차감액".format(period), WON.format(sample) + " 원")

        st.divider()
        if st.button("입력한 값 저장", type="primary",
                     help="설정.xlsx 에 저장됩니다. 다음에 열면 그대로 나오고, 배치(실행.bat)도 같은 값을 씁니다."):
            try:
                if USE_DB:
                    db.save_cards(period, cards_in)
                    db.save_samples(smp_year)
                    db.save_channels(channels)
                    db.save_setting("period", period)
                    st.success("Supabase 에 저장했습니다.")
                else:
                    save_config(period, def_sources, def_years, ch_df, cards_in, smp_year,
                                def_cards_all, def_samples_all)
                    st.success("저장했습니다 — 입력/설정.xlsx")
                read_config_defaults.clear()
            except PermissionError:
                st.error("설정.xlsx 가 엑셀에서 열려 있습니다. 닫고 다시 시도하세요.")
            except Exception as e:
                st.error("저장 실패: {}".format(e))

    st.divider()

    def take_snapshot():
        """계산에 필요한 모든 것을 한 덩어리로 담아둔다.
           결과 화면에서는 입력 위젯이 그려지지 않으므로, 재계산 때 이 값을 쓴다."""
        files = {}
        for ch in channels:
            up = uploads.get(ch["name"])
            if up is not None:
                files[ch["name"]] = (up.name, up.getvalue())
        return {
            "channels": [dict(c) for c in channels], "files": files,
            "cost": cost, "fee": fee, "origin": origin, "conflicts": conflicts,
            "override": override, "fee_override": fee_override,
            "cards": cards, "card_total": card_total,
            "sample_items": sample_items, "sample": sample, "period": period,
        }


    ready = cost is not None and (
        any(uploads.get(ch["name"]) is not None for ch in channels)
        or any(ch["name"] in _prev for ch in channels))
    if st.button("이익률 계산", type="primary", disabled=not ready, width="stretch"):
        snap = take_snapshot()
        st.session_state["calc_snapshot"] = snap
        ok, errs = run_calc(snap)
        if ok:
            go("result")                                 # 결과 화면으로 이동
            st.rerun()
        else:
            st.error("계산할 파일이 없습니다. 아래 사유를 확인하세요.")
            for e in errs:
                st.warning(e)

    _prev_files = (st.session_state.get("calc_snapshot") or {}).get("files", {})
    _now_up = any(uploads.get(ch["name"]) is not None for ch in channels)
    if _prev_files and not _now_up:
        st.info("직전 계산에 쓴 매출 파일 **{}개** 가 남아 있습니다 — {}\n\n"
                "그대로 다시 계산하거나, 새 파일을 올려 덮어쓸 수 있습니다."
                .format(len(_prev_files), " · ".join(_prev_files)))
    elif not ready:
        if cost is None:
            st.info("먼저 사이드바에서 **이익률 마스터** 를 지정하세요.")
        else:
            st.info("**① 이카운트 매출 파일** 에 파일을 올리면 계산할 수 있습니다.")

    if st.session_state.get("out"):
        st.caption("")
        if st.button("↩ 지난 계산 결과 다시 보기", width="stretch"):
            go("result")
            st.rerun()

    _saved_in = list_reports()
    if _saved_in:
        with st.expander("월별 보관함 — 지난 달 결과 내려받기 ({}개월)".format(len(_saved_in))):
            for _p, _fn, _sm, _at in _saved_in:
                _c = st.columns([1.2, 3.4, 1.4])
                _c[0].markdown("**{}**".format(_p))
                _c[1].markdown("<span style='color:#64748B;font-size:.82rem'>{}</span>"
                               .format(_sm or "-"), unsafe_allow_html=True)
                _got = load_report(_p)
                if _got:
                    _c[2].download_button("내려받기", data=_got[1], file_name=_got[0],
                                          mime="application/vnd.openxmlformats-officedocument."
                                               "spreadsheetml.sheet",
                                          key="dli_%s" % _p, width="stretch")

# ---------------------------------------------------------------- 결과
out = st.session_state.get("out")
if out:
    results, total, errs = out["results"], out["total"], out["errs"]
    for e in errs:
        st.warning(e)

    st.header("{} 결과".format(out["period"]))
    m = st.columns(6)
    m[0].metric("총매출 (배송 포함)", WON.format(total["gross"]) + " 원")
    m[1].metric("상품매출", WON.format(total["sales"]) + " 원")
    m[2].metric("배송매출", WON.format(total["delivery"]) + " 원",
                delta="이익 계산 제외", delta_color="off")
    m[3].metric("총이익액", WON.format(total["profit"]) + " 원")
    m[4].metric("이익률", "{:.2f} %".format(total["margin"] * 100))
    miss = E.missing_rows(results)
    m[5].metric("원가 미매칭", "{}건".format(len(miss)),
                delta="판매액 " + WON.format(sum(x[4] for x in miss)) + "원",
                delta_color="inverse" if miss else "off")

    # ---- 채널별 한눈에 보기
    st.write("")
    st.markdown("##### 채널별")
    ch_cols = st.columns(len(results) or 1)
    for col, res in zip(ch_cols, results):
        rate_ = res["profit"] / res["sales"] if res["sales"] else 0.0
        tone = "#10B981" if rate_ >= 0.25 else ("#6366F1" if rate_ >= 0.15 else "#F59E0B")
        cuts = []
        if res["fee"]:
            cuts.append(("셀러수수료", res["fee"]))
        if res["card"]:
            cuts.append(("카드수수료", res["card"]))
        if res["sample"]:
            cuts.append(("샘플비용", res["sample"]))
        cut_html = "".join(
            "<div style='display:flex;justify-content:space-between;font-size:.76rem;"
            "color:#94A3B8;margin-top:.15rem'><span>− {}</span><span>{}</span></div>"
            .format(n, WON.format(v)) for n, v in cuts)
        col.markdown(
            "<div style='background:#FFFFFF;border:1px solid #E2E8F0;border-radius:12px;"
            "padding:1.05rem 1.15rem;box-shadow:0 1px 2px rgba(15,23,42,.04)'>"
            "<div style='font-size:.86rem;font-weight:700;color:#0F172A'>{name}</div>"
            "<div style='font-size:1.75rem;font-weight:700;color:{tone};"
            "letter-spacing:-.03em;line-height:1.3;margin:.15rem 0 .5rem'>{rate:.2f}%</div>"
            "<div style='display:flex;justify-content:space-between;font-size:.8rem;"
            "color:#64748B'><span>상품매출</span><span style='color:#0F172A'>{sales}</span></div>"
            "<div style='display:flex;justify-content:space-between;font-size:.8rem;"
            "color:#64748B;margin-top:.15rem'><span>상품원가</span><span>{cost}</span></div>"
            "{cuts}"
            "<div style='border-top:1px solid #E2E8F0;margin:.5rem 0 .35rem'></div>"
            "<div style='display:flex;justify-content:space-between;font-size:.86rem;"
            "font-weight:700;color:#0F172A'><span>이익액</span><span>{profit}</span></div>"
            "<div style='display:flex;justify-content:space-between;font-size:.75rem;"
            "color:#94A3B8;margin-top:.3rem'><span>배송매출(제외)</span><span>{deli}</span></div>"
            "</div>".format(
                name=res["ch"]["name"], tone=tone, rate=rate_ * 100,
                sales=WON.format(res["sales"]), cost=WON.format(res["cost"]),
                cuts=cut_html, profit=WON.format(res["profit"]),
                deli=WON.format(res["delivery_sales"])),
            unsafe_allow_html=True)
    st.caption("이익률 = 이익액 ÷ 상품매출 · 배송은 매출=원가로 처리해 이익에서 제외")
    st.write("")

    st.download_button(
        "결과 엑셀 다운로드",
        data=out["xlsx"],
        file_name="{} 이익률.xlsx".format(out["period"]),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary", width="stretch")

    _saved = list_reports()
    if _saved:
        with st.expander("월별 보관함 — 지난 달 결과 내려받기 ({}개월)".format(len(_saved))):
            st.caption("계산할 때마다 그 달의 결과 엑셀이 자동으로 보관됩니다. "
                       "다른 PC에서 접속해도 같은 목록이 보입니다.")
            for _p, _fn, _sm, _at in _saved:
                _c = st.columns([1.2, 3, 1.6, 1.4])
                _c[0].markdown("**{}**".format(_p))
                _c[1].markdown("<span style='color:#64748B;font-size:.82rem'>{}</span>"
                               .format(_sm or "-"), unsafe_allow_html=True)
                _c[2].markdown("<span style='color:#94A3B8;font-size:.78rem'>{}</span>"
                               .format(_at.strftime("%Y-%m-%d %H:%M") if _at else ""),
                               unsafe_allow_html=True)
                _got = load_report(_p)
                if _got:
                    _c[3].download_button("내려받기", data=_got[1], file_name=_got[0],
                                          mime="application/vnd.openxmlformats-officedocument."
                                               "spreadsheetml.sheet",
                                          key="dl_%s" % _p, width="stretch")

    summary = pd.DataFrame([{
        "구분": r["ch"]["name"], "상품매출": r["sales"], "상품원가": r["cost"],
        "셀러수수료": r["fee"], "카드수수료": r["card"], "샘플비용": r["sample"],
        "이익액": r["profit"],
        "이익률": r["profit"] / r["sales"] if r["sales"] else 0.0,
        "배송매출": r["delivery_sales"],
    } for r in results] + [{
        "구분": "합계", "상품매출": total["sales"], "상품원가": total["cost"],
        "셀러수수료": total["fee"], "카드수수료": total["card"], "샘플비용": total["sample"],
        "이익액": total["profit"],
        "이익률": total["profit"] / total["sales"] if total["sales"] else 0.0,
        "배송매출": total["delivery"],
    }])
    money_cols = ["상품매출", "상품원가", "셀러수수료", "카드수수료", "샘플비용", "이익액", "배송매출"]
    st.dataframe(
        summary, width="stretch", hide_index=True,
        column_config={**{c: st.column_config.NumberColumn(format="%d") for c in money_cols},
                       "이익률": st.column_config.NumberColumn(format="percent")})
    st.caption("* 배송 이익 제외 — 배송매출은 원가와 같게 처리하여 이익 0")

    # ---- 월별 추이
    hist = load_history()
    if hist:
        import altair as alt
        hdf = pd.DataFrame(hist)
        hdf["이익률"] = hdf.apply(
            lambda r: (r["profit"] / r["sales"]) if r["sales"] else 0.0, axis=1)
        months = sorted(hdf["period"].unique())

        if len(months) >= 2:
            st.write("")
            st.markdown("##### 월별 추이")

            INK, MUTED, LINE, SURF = "#0F172A", "#64748B", "#E2E8F0", "#FFFFFF"
            # 검증된 카테고리 순서 (blue → orange → aqua → yellow) + 전체는 강조색
            CH_ORDER = [r["ch"]["name"] for r in results]
            CH_HUES = ["#2a78d6", "#eb6834", "#1baf7a", "#eda100",
                       "#e87ba4", "#008300", "#4a3aa7", "#e34948"][:len(CH_ORDER)]
            AX = alt.Axis(labelColor=MUTED, titleColor=MUTED, tickColor=LINE,
                          domainColor=LINE, labelFontSize=11, titleFontSize=11,
                          grid=True, gridColor=LINE, gridOpacity=.7)

            g1, g2 = st.columns(2)

            # ① 전체 이익률 — 단일 계열이라 범례 없이 제목이 이름을 대신한다
            tot = hdf[hdf["channel"] == "전체"].sort_values("period")
            base = alt.Chart(tot).encode(
                x=alt.X("period:O", title=None, axis=AX),
                y=alt.Y("이익률:Q", title="이익률", axis=alt.Axis(format="%", **{
                    k: v for k, v in AX.to_dict().items() if k != "format"})),
                tooltip=[alt.Tooltip("period:O", title="기준월"),
                         alt.Tooltip("이익률:Q", format=".2%"),
                         alt.Tooltip("sales:Q", title="상품매출", format=",.0f"),
                         alt.Tooltip("profit:Q", title="이익액", format=",.0f")])
            g1.markdown("**전체 이익률**")
            g1.altair_chart(
                (base.mark_line(color="#6366F1", strokeWidth=2, point=alt.OverlayMarkDef(
                    color="#6366F1", size=60, filled=True))
                 + base.mark_text(dy=-14, fontSize=11, color=INK).encode(
                     text=alt.Text("이익률:Q", format=".1%"))
                 ).properties(height=260, background=SURF)
                .configure_view(strokeWidth=0), use_container_width=True)

            # ② 채널별 이익률 — 계열이 곧 주제이므로 categorical
            ch = hdf[hdf["channel"] != "전체"].sort_values(["period", "channel"])
            g2.markdown("**채널별 이익률**")
            g2.altair_chart(
                alt.Chart(ch).mark_line(strokeWidth=2, point=alt.OverlayMarkDef(
                    size=45, filled=True)).encode(
                    x=alt.X("period:O", title=None, axis=AX),
                    y=alt.Y("이익률:Q", title="이익률", axis=alt.Axis(format="%", **{
                        k: v for k, v in AX.to_dict().items() if k != "format"})),
                    color=alt.Color("channel:N", title=None,
                                    scale=alt.Scale(domain=CH_ORDER, range=CH_HUES),
                                    legend=alt.Legend(orient="top", labelColor=MUTED,
                                                      symbolStrokeWidth=2, labelFontSize=11)),
                    tooltip=[alt.Tooltip("period:O", title="기준월"),
                             alt.Tooltip("channel:N", title="채널"),
                             alt.Tooltip("이익률:Q", format=".2%"),
                             alt.Tooltip("profit:Q", title="이익액", format=",.0f")]
                ).properties(height=260, background=SURF)
                .configure_view(strokeWidth=0), use_container_width=True)

            # ③ 매출과 이익액 — 같은 단위(원)라 한 축에 둘 다 올린다 (이중축 금지)
            st.markdown("**월별 상품매출 · 이익액**")
            bars = tot.melt(id_vars="period", value_vars=["sales", "profit"],
                            var_name="구분", value_name="금액")
            bars["구분"] = bars["구분"].map({"sales": "상품매출", "profit": "이익액"})
            st.altair_chart(
                alt.Chart(bars).mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4,
                                         strokeWidth=2, stroke=SURF).encode(
                    x=alt.X("period:O", title=None, axis=AX),
                    xOffset=alt.XOffset("구분:N"),
                    y=alt.Y("금액:Q", title="원", axis=AX),
                    color=alt.Color("구분:N", title=None,
                                    scale=alt.Scale(domain=["상품매출", "이익액"],
                                                    range=["#2a78d6", "#eb6834"]),
                                    legend=alt.Legend(orient="top", labelColor=MUTED,
                                                      labelFontSize=11)),
                    tooltip=[alt.Tooltip("period:O", title="기준월"),
                             alt.Tooltip("구분:N"),
                             alt.Tooltip("금액:Q", format=",.0f")]
                ).properties(height=270, background=SURF)
                .configure_view(strokeWidth=0), use_container_width=True)

            with st.expander("월별 표로 보기"):
                piv = hdf.pivot_table(index="period", columns="channel",
                                      values="이익률", aggfunc="first").reset_index()
                piv = piv.rename(columns={"period": "기준월"})
                st.dataframe(piv, width="stretch", hide_index=True,
                             column_config={c: st.column_config.NumberColumn(format="percent")
                                            for c in piv.columns if c != "기준월"})
        else:
            st.caption("월별 추이 그래프는 **2개월 이상** 계산하면 나타납니다. "
                       "(현재 {}개월 기록)".format(len(months)))

    tabs = st.tabs(["배송비", "미매칭", "셀러수수료 기본율", "보정적용내역", "중복코드 점검"]
                   + [r["ch"]["name"] for r in results])

    with tabs[0]:
        st.info("배송비는 **매출 = 원가**로 처리해 이익에서 제외합니다 (배송 제외 이익률). "
                "아래는 이번 달 배송 매출 내역입니다.")
        drows = []
        for res in results:
            for d in res["delivery"]:
                drows.append({"채널": res["ch"]["name"], "품목그룹3코드": d["group_code"],
                              "품목명[규격]": d["name"] or d["code_raw"],
                              "수량": d["qty"], "판매액": d["amount"]})
        if not drows:
            st.warning("배송비 행을 찾지 못했습니다. "
                       "이카운트 파일에 `품목그룹3코드` 열이 있고 배송비가 `Z0001` 인지 확인하세요.")
        else:
            dd = pd.DataFrame(drows)
            bych = dd.groupby("채널", as_index=False)[["수량", "판매액"]].sum()
            bych.loc[len(bych)] = ["합계", bych["수량"].sum(), bych["판매액"].sum()]
            c1, c2 = st.columns([1, 2])
            with c1:
                st.markdown("**채널별**")
                st.dataframe(bych, width="stretch", hide_index=True,
                             column_config={k: st.column_config.NumberColumn(format="%d")
                                            for k in ["수량", "판매액"]})
            with c2:
                st.markdown("**품목별**")
                st.dataframe(dd.sort_values("판매액", ascending=False),
                             width="stretch", hide_index=True,
                             column_config={k: st.column_config.NumberColumn(format="%d")
                                            for k in ["수량", "판매액"]})

    with tabs[1]:
        manual = st.session_state.setdefault("manual_cost", {})
        used = out.get("manual_used", {})
        if used:
            st.success("직접 입력한 원가 {}건이 이번 계산에 반영되었습니다.".format(len(used)))
        if miss:
            st.error("원가를 못 찾은 {}건 — 원가가 `0` 으로 잡혀 **이익이 과대계상**됩니다. "
                     "아래 `원가(직접입력)` 칸에 넣고 [입력한 원가로 다시 계산] 을 누르세요."
                     .format(len(miss)))
            miss_df = st.data_editor(
                pd.DataFrame([{"채널": c, "상품코드": raw, "품목명[규격]": nm,
                               "수량": q, "판매액": a,
                               "원가(직접입력)": float(manual.get(_k, 0.0) or 0.0)}
                              for _k, raw, nm, q, a, c in miss]),
                width="stretch", hide_index=True, key="miss_editor",
                disabled=["채널", "상품코드", "품목명[규격]", "수량", "판매액"],
                column_config={
                    "판매액": st.column_config.NumberColumn(format="%d"),
                    "수량": st.column_config.NumberColumn(format="%d"),
                    "원가(직접입력)": st.column_config.NumberColumn(
                        format="%d", min_value=0.0, step=100.0,
                        help="1개당 원가를 넣으세요. 수량은 자동으로 곱해집니다."),
                })
            typed = {}
            for _i, r in miss_df.iterrows():
                v = E.num(r["원가(직접입력)"])
                if v > 0:
                    typed[E.code_key(r["상품코드"])] = v

            b1, b2 = st.columns([1, 1])
            if b1.button("입력한 원가로 다시 계산", type="primary",
                         disabled=not typed, width="stretch"):
                st.session_state["manual_cost"] = {**manual, **typed}
                ok, _e = run_calc()
                if ok:
                    st.rerun()
            if b2.button(("Supabase 에 저장 (매달 자동 적용)" if USE_DB
                          else "원가보정.xlsx 에 저장 (매달 자동 적용)"),
                         disabled=not typed, width="stretch",
                         help="다음 달부터는 입력하지 않아도 이 원가가 쓰입니다"):
                try:
                    names = {E.code_key(r["상품코드"]): (str(r["상품코드"]), str(r["품목명[규격]"]))
                             for _i, r in miss_df.iterrows()}
                    if USE_DB:
                        n = db.upsert_overrides({k: (v, False) for k, v in typed.items()},
                                                names, note="화면에서 직접 입력")
                        where = "Supabase"
                    else:
                        n = append_overrides(typed, names)
                        where = "원가보정.xlsx"
                    st.session_state["manual_cost"] = {**manual, **typed}
                    load_override_cached.clear()
                    st.success("{} 에 {}건 저장했습니다. 다시 계산합니다.".format(where, n))
                    ok, _e = run_calc()
                    if ok:
                        st.rerun()
                except PermissionError:
                    st.error("원가보정.xlsx 가 엑셀에서 열려 있습니다. 닫고 다시 시도하세요.")
        else:
            st.success("모든 상품의 원가를 찾았습니다.")
        if manual:
            with st.expander("직접 입력한 원가 {}건 — 되돌리기".format(len(manual))):
                st.dataframe(pd.DataFrame([{"상품코드": k, "원가": v}
                                           for k, v in manual.items()]),
                             width="stretch", hide_index=True,
                             column_config={"원가": st.column_config.NumberColumn(format="%d")})
                if st.button("직접 입력한 원가 모두 지우기"):
                    st.session_state["manual_cost"] = {}
                    run_calc()
                    st.rerun()

    with tabs[2]:
        rows = E.fee_default_rows(results)
        if rows:
            st.warning("기준 파일에 요율이 없어 기본율로 계산된 {}건 "
                       "(합계 {} 원)".format(len(rows), WON.format(sum(x[5] for x in rows))))
            st.dataframe(pd.DataFrame(
                [{"채널": a, "상품코드": b, "품목명[규격]": c, "판매액": d,
                  "적용요율": e, "수수료": f} for a, b, c, d, e, f in rows]),
                width="stretch", hide_index=True,
                column_config={"판매액": st.column_config.NumberColumn(format="%d"),
                               "수수료": st.column_config.NumberColumn(format="%d"),
                               "적용요율": st.column_config.NumberColumn(format="percent")})
        else:
            st.success("모든 상품이 기준 파일에서 매칭되었습니다.")

    with tabs[3]:
        rows = E.applied_rows(results)
        if rows:
            over = [r for r in rows if r[6] is not None]
            if over:
                st.warning("기준 파일에 값이 있는데도 보정값으로 덮어쓴 {}건이 있습니다. "
                           "원가 증감 {} 원 — 어느 쪽이 맞는지 확인하세요.".format(
                               len(over), WON.format(sum(r[6] for r in over))))
            st.dataframe(pd.DataFrame(
                [{"채널": a, "상품코드": b, "품목명[규격]": c, "수량": d,
                  "기준 원가": e, "적용 원가": f, "차이(원가합)": g}
                 for a, b, c, d, e, f, g in rows]),
                width="stretch", hide_index=True,
                column_config={k: st.column_config.NumberColumn(format="%d")
                               for k in ["수량", "기준 원가", "적용 원가", "차이(원가합)"]})
        else:
            st.info("적용된 보정이 없습니다.")

    with tabs[4]:
        if not conflicts:
            st.success("중복 등록된 상품코드가 없습니다.")
        else:
            sold = {}
            for res in results:
                for g in res["goods"]:
                    a = sold.setdefault(g["code"], {"qty": 0.0, "amt": 0.0, "name": g["name"]})
                    a["qty"] += g["qty"]
                    a["amt"] += g["amount"]
            hit = [c for c in conflicts if c["code"] in sold]
            risk = sum(sold[c["code"]]["qty"] * c["spread"] for c in hit)
            amt = sum(sold[c["code"]]["amt"] for c in hit)

            st.warning(
                "같은 상품코드가 **서로 다른 원가**로 여러 번 등록되어 있습니다 — "
                "**기재날짜가 가장 최근인 값**을 사용합니다.\n\n"
                "- 중복 코드 **{:,}개** (이번 달 실제 판매된 것 **{:,}개**, 판매액 {} 원)\n"
                "- 어느 값을 쓰느냐에 따라 원가가 최대 **{} 원**까지 달라질 수 있습니다"
                .format(len(conflicts), len(hit), WON.format(amt), WON.format(risk)))

            only_sold = st.checkbox("이번 달 판매된 것만 보기", value=True, key="conf_sold")
            show = hit if only_sold else conflicts
            show = sorted(show, key=lambda c: -(sold.get(c["code"], {}).get("amt", 0)
                                                if only_sold else c["spread"]))
            if not show:
                st.info("이번 달 판매된 상품 중에는 중복 코드가 없습니다.")
            else:
                st.caption("★ 가 실제로 사용된 값입니다. "
                           + ("판매액이 큰 순서" if only_sold else "금액 차이가 큰 순서"))
                rows = []
                for c in show:
                    s = sold.get(c["code"], {})
                    for r in c["rows"]:
                        rows.append({
                            "상품코드": c["code"],
                            "사용": "★" if r.get("used") else "",
                            "기재날짜": r["date_txt"],
                            "시트": r["sheet"],
                            "상품명": r["name"],
                            "옵션명": r["opt"],
                            "원가": r["value"],
                            "최고-최저": c["spread"],
                            "이번달 판매액": s.get("amt", 0),
                            "이번달 수량": s.get("qty", 0),
                        })
                st.dataframe(
                    pd.DataFrame(rows), width="stretch", hide_index=True, height=520,
                    column_config={k: st.column_config.NumberColumn(format="%d")
                                   for k in ["원가", "최고-최저", "이번달 판매액", "이번달 수량"]})
                st.caption("같은 코드에 상품명·옵션명이 다르게 붙어 있다면, "
                           "마진율표에서 `공급처옵션` 코드가 여러 옵션에 공용으로 쓰인 것입니다. "
                           "정확한 원가가 필요하면 옵션별로 코드를 나누거나 "
                           "`원가보정.xlsx` 에 해당 코드를 직접 지정하세요.")

    for i, res in enumerate(results, start=5):
        with tabs[i]:
            ch = res["ch"]
            rows = []
            for g in res["goods"]:
                row = {"품목그룹3": g["group_name"], "상품코드": g["code_raw"],
                       "품목명[규격]": g["name"], "수량": g["qty"], "판매액": g["amount"],
                       "원가": g["cost"], "원가합": g["cost_sum"], "원가출처": g["cost_src"]}
                if ch["seller"]:
                    row["셀러수수료율"] = g["fee_rate"]
                    row["요율출처"] = g["fee_src"]
                    row["셀러수수료"] = g["fee_sum"]
                p = g["amount"] - g["cost_sum"] - g["fee_sum"]
                row["이익액"] = p
                row["이익률"] = p / g["amount"] if g["amount"] else 0.0
                rows.append(row)
            df = pd.DataFrame(rows)
            only_bad = st.checkbox("원가 미매칭만 보기", key="flt_" + ch["name"])
            if only_bad:
                df = df[df["원가출처"] == "미매칭"]
            fmt = {c: st.column_config.NumberColumn(format="%d")
                   for c in ["수량", "판매액", "원가", "원가합", "셀러수수료", "이익액"] if c in df.columns}
            fmt["이익률"] = st.column_config.NumberColumn(format="percent")
            if "셀러수수료율" in df.columns:
                fmt["셀러수수료율"] = st.column_config.NumberColumn(format="percent")
            st.dataframe(df, width="stretch", hide_index=True,
                         column_config=fmt, height=460)
            if res["delivery"]:
                st.caption("배송비 (이익 계산 제외) — {} 원".format(
                    WON.format(res["delivery_sales"])))
                st.dataframe(pd.DataFrame(
                    [{"품목명[규격]": d["name"], "수량": d["qty"], "판매액": d["amount"]}
                     for d in res["delivery"]]),
                    width="stretch", hide_index=True,
                    column_config={"수량": st.column_config.NumberColumn(format="%d"),
                                   "판매액": st.column_config.NumberColumn(format="%d")})

    if not USE_DB and st.button("결과 폴더에도 저장"):
        OUT_DIR.mkdir(exist_ok=True)
        p = OUT_DIR / "{} 이익률.xlsx".format(out["period"])
        try:
            p.write_bytes(out["xlsx"])
            st.success("저장했습니다 — 결과/{}".format(p.name))
        except PermissionError:
            st.error("'{}' 가 엑셀에서 열려 있습니다. 닫고 다시 시도하세요.".format(p.name))
