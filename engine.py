# -*- coding: utf-8 -*-
"""
이익률 계산 엔진 — 파일 경로에 의존하지 않는 순수 계산 로직.

CLI(이익률계산.py) 와 웹앱(앱.py) 이 이 모듈을 공유한다.
모든 입력은 경로(str/Path) 또는 파일객체(BytesIO 등) 를 받는다.

계산식
  일반 채널 : 이익 = 판매액 - (원가 x 수량)
  스룩      : 이익 = 판매액 - (원가 x 수량) - 셀러수수료 - 카드수수료
  배송비 행(품목그룹3코드 Z0001)은 매출=원가로 처리하여 이익 0
  샘플비용/카드수수료는 지정한 채널의 이익에서 총액으로 차감
"""

import re
import unicodedata
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DELIVERY_GROUP_CODES = {"Z0001"}
DELIVERY_GROUP_NAMES = {"배송비"}

# 이카운트 다운로드에서 찾을 열 이름 (동의어 허용)
COLUMN_ALIASES = {
    "group_code": ["품목그룹3코드", "품목그룹코드", "그룹3코드", "품목그룹1코드", "품목그룹2코드"],
    "group_name": ["품목그룹3", "품목그룹", "그룹3", "품목그룹1", "품목그룹2"],
    "code": ["품목명[규격]코드", "품목명(규격)코드", "품목명규격코드", "품목[규격]코드",
             "품목코드", "관리옵션코드", "상품코드", "자체품목코드", "옵션코드"],
    "name": ["품목명[규격]", "품목명(규격)", "품목명규격", "품목[규격]",
             "품목명", "상품명", "품목"],
    "qty": ["수량", "판매수량", "출고수량", "매출수량"],
    "amount": ["판매액", "매출액", "판매금액", "매출금액", "공급가액", "합계금액", "금액"],
}

DEFAULT_CHANNELS = [
    {"name": "유통", "file": "유통.xlsx", "seller": False, "card": False, "sample": True, "fee_base": 0.0},
    {"name": "유통B2B", "file": "유통B2B.xlsx", "seller": False, "card": False, "sample": False, "fee_base": 0.0},
    {"name": "유통_입점몰", "file": "유통_입점몰.xlsx", "seller": False, "card": False, "sample": False, "fee_base": 0.0},
    {"name": "스룩", "file": "스룩.xlsx", "seller": True, "card": True, "sample": False, "fee_base": 0.15},
]

DEFAULT_CARDS = [("신용카드", 0.0275), ("무통장입금", 0.0), ("페이코", 0.0275), ("토스", 0.0275)]


# ---------------------------------------------------------------- 유틸
def norm(v):
    """헤더/코드 비교용 정규화: 공백·전각문자 제거"""
    if v is None:
        return ""
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(v)))


def code_key(v):
    """상품코드 매칭 키 (앞뒤 공백 제거, 대소문자 무시)"""
    if v is None:
        return ""
    s = unicodedata.normalize("NFKC", str(v)).strip()
    if s.endswith(".0") and s[:-2].isdigit():   # 숫자코드가 실수로 읽힌 경우
        s = s[:-2]
    return s.upper()


def num(v):
    """숫자 변환 (콤마·원·% 등 제거). 실패 시 0"""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d.\-]", "", str(v))
    if s in ("", "-", "."):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def rate(v):
    """수수료율 변환. 15 / '15%' / 0.15 를 모두 0.15 로"""
    if v is None:
        return None
    if isinstance(v, str) and "%" in v:
        return num(v) / 100.0
    x = num(v)
    return x / 100.0 if x > 1 else x


def is_yes(v):
    return norm(v).upper() in {"O", "Y", "YES", "TRUE", "1", "적용", "예"}


_HEADER_WORDS = {norm(a) for al in COLUMN_ALIASES.values() for a in al}
_SKIP_WORDS = {"합계", "소계", "총계"}


def is_junk(code_raw, name):
    """반복 삽입된 헤더행·합계행 걸러내기"""
    return (norm(code_raw) in _HEADER_WORDS
            or norm(name) in _SKIP_WORDS
            or norm(code_raw) in _SKIP_WORDS)


# ---------------------------------------------------------------- 스룩페이 매출통계
# 스룩페이 → 매출/정산 → [엑셀 내려받기] 로 받은 파일에서 결제수단별 금액을 읽는다.
#   시트 '매출내역 기간 통계' :  결제수단 | ... | 정상금액 | ... | 수수료
#   정상금액 = 결제합계 - 취소합계   (7월까지 손으로 옮겨 적던 그 값)
#   수수료   = 스룩페이가 건별로 계산한 실제 수수료 (요율 x 금액보다 정확)
SROOK_SHEETS = ("매출내역 기간 통계", "매출내역기간통계")
SROOK_DAILY = ("일별 매출 리스트(합계)", "일별매출리스트(합계)")


def read_srookpay(src):
    """스룩페이 매출통계 엑셀 → dict
       {rows: [(결제수단, 정상금액, 수수료, 추정요율)], total_amount, total_fee,
        period_from, period_to}"""
    wb = open_wb(src, read_only=True, data_only=True)
    try:
        sheet = None
        for ws in wb.worksheets:
            if norm(ws.title) in {norm(x) for x in SROOK_SHEETS}:
                sheet = list(ws.iter_rows(values_only=True))
                break
        if sheet is None:                      # 시트 이름이 달라도 헤더로 찾아본다
            for ws in wb.worksheets:
                rows = list(ws.iter_rows(values_only=True))
                if rows and any(norm(v) == "결제수단" for v in rows[0] if v):
                    sheet = rows
                    break
        if sheet is None:
            raise ValueError(
                "'매출내역 기간 통계' 시트를 찾지 못했습니다.\n"
                "스룩페이 → 매출/정산 화면에서 [엑셀 내려받기] 로 받은 파일이 맞는지 확인하세요.")

        hdr = {norm(v): i for i, v in enumerate(sheet[0]) if v}
        need = ("결제수단", "정상금액")
        miss = [n for n in need if n not in hdr]
        if miss:
            raise ValueError("이 파일에 [{}] 열이 없습니다. 열 이름: {}".format(
                " · ".join(miss), " | ".join(list(hdr)[:12])))

        i_m, i_a = hdr["결제수단"], hdr["정상금액"]
        i_f = hdr.get("수수료")
        out, t_amt, t_fee = [], 0.0, 0.0
        for r in sheet[1:]:
            m = r[i_m] if i_m < len(r) else None
            if not m or not str(m).strip():
                continue
            name = str(m).strip()
            amt = num(r[i_a]) if i_a < len(r) else 0.0
            fee = num(r[i_f]) if (i_f is not None and i_f < len(r)) else 0.0
            if norm(name) in ("소계", "합계", "총계"):     # 소계는 검산용으로만
                t_amt, t_fee = amt, fee
                continue
            out.append((name, amt, fee, (fee / amt) if amt else 0.0))
        if not out:
            raise ValueError("결제수단 행을 읽지 못했습니다.")
        if not t_amt:
            t_amt = sum(x[1] for x in out)
            t_fee = sum(x[2] for x in out)

        # 조회 기간 — 일별 시트의 첫/마지막 날짜
        p_from = p_to = ""
        for ws in wb.worksheets:
            if norm(ws.title) in {norm(x) for x in SROOK_DAILY}:
                days = []
                for r in list(ws.iter_rows(values_only=True))[1:]:
                    v = r[0] if r else None
                    if not v:
                        continue
                    m2 = re.search(r"(20\d{2}-\d{2}-\d{2})", str(v))
                    if m2:
                        days.append(m2.group(1))
                if days:
                    p_from, p_to = min(days), max(days)
                break
    finally:
        wb.close()
    return {"rows": out, "total_amount": t_amt, "total_fee": t_fee,
            "period_from": p_from, "period_to": p_to}


def is_delivery(r):
    return (r["group_code"].upper() in DELIVERY_GROUP_CODES
            or norm(r["group_name"]) in {norm(x) for x in DELIVERY_GROUP_NAMES})


def open_wb(src, **kw):
    """경로 또는 파일객체 모두 처리"""
    if hasattr(src, "seek"):
        src.seek(0)
    try:
        wb = openpyxl.load_workbook(src if hasattr(src, "read") else str(src), **kw)
        # 이카운트 등 일부 프로그램은 시트 크기를 '1행 1열'로 잘못 적어둔다.
        # 읽기전용 모드에서는 그 값을 그대로 믿기 때문에 데이터를 놓친다 → 다시 재도록 초기화
        for ws in wb.worksheets:
            if hasattr(ws, "reset_dimensions"):
                ws.reset_dimensions()
        return wb
    except Exception as e:
        msg = str(e)
        if "zip" in msg.lower() or "not a zip" in msg.lower():
            raise ValueError(
                "엑셀 파일로 열리지 않습니다. 확장자는 .xlsx 인데 실제로는 구형 .xls 이거나 "
                "웹 화면을 저장한 파일일 수 있습니다.\n"
                "엑셀에서 열어 [다른 이름으로 저장] → 'Excel 통합 문서(*.xlsx)' 로 "
                "다시 저장한 뒤 올려주세요.")
        raise


def find_header(ws, need, scan_rows=60, rows=None):
    """헤더 행을 찾아 (행번호, {필드: 열번호}) 반환. 하나라도 없으면 (None, None).
       시트 크기 정보가 없거나 틀려도 되도록 iter_rows 로 훑는다.
       rows 를 주면 이미 읽어둔 값 목록을 쓴다 (같은 시트를 두 번 훑지 않도록)."""
    src = rows[:scan_rows] if rows is not None else ws.iter_rows(
        max_row=scan_rows, values_only=True)
    for r, row in enumerate(src, start=1):
        seen = {}
        for c, v in enumerate(row, start=1):
            k = norm(v)
            if k and k not in seen:
                seen[k] = c
        found = {}
        for field, aliases in need.items():
            for a in aliases:
                if norm(a) in seen:
                    found[field] = seen[norm(a)]
                    break
        if all(f in found for f in need):
            return r, found
    return None, None


def _pick(row, col, field):
    if field not in col:
        return None
    i = col[field] - 1
    return row[i] if i < len(row) else None


# ---------------------------------------------------------------- 마스터
def load_master(src):
    """원가·셀러수수료 조회표 생성. VLOOKUP 과 동일하게 '처음 나온 값' 우선.
       반환: (cost, fee, conflicts)"""
    wb = open_wb(src, read_only=True, data_only=True)
    try:
        if "원가리스트_최종" not in wb.sheetnames:
            raise ValueError("마스터에 '원가리스트_최종' 시트가 없습니다.")
        if "셀러수수료_최종" not in wb.sheetnames:
            raise ValueError("마스터에 '셀러수수료_최종' 시트가 없습니다.")

        cost, cost_dup = {}, defaultdict(set)
        ws = wb["원가리스트_최종"]
        hr, col = find_header(ws, {"code": ["상품코드", "관리옵션코드"], "cost": ["원가"]})
        if hr is None:
            raise ValueError("'원가리스트_최종' 에서 [상품코드/원가] 열을 찾지 못했습니다.")
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            k = code_key(_pick(row, col, "code"))
            if not k:
                continue
            v = num(_pick(row, col, "cost"))
            cost_dup[k].add(v)
            cost.setdefault(k, v)

        fee, fee_dup = {}, defaultdict(set)
        ws = wb["셀러수수료_최종"]
        hr, col = find_header(ws, {"code": ["상품코드"], "fee": ["셀러수수료"]})
        if hr is None:
            raise ValueError("'셀러수수료_최종' 에서 [상품코드/셀러수수료] 열을 찾지 못했습니다.")
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            k = code_key(_pick(row, col, "code"))
            if not k:
                continue
            v = rate(_pick(row, col, "fee"))
            if v is None:
                continue
            fee_dup[k].add(v)
            fee.setdefault(k, v)
    finally:
        wb.close()

    # 구마스터에는 시트·날짜 정보가 없어 값 목록만 표시한다
    conflicts = []
    for k, v in cost_dup.items():
        if len(v) > 1:
            vals = sorted(v)
            conflicts.append({
                "code": k, "used": cost[k], "used_date": "", "values": vals,
                "spread": vals[-1] - vals[0],
                "rows": [{"sheet": "원가리스트_최종", "row": 0, "date": _FAR_PAST,
                          "date_txt": "", "name": "", "opt": "", "value": x,
                          "used": x == cost[k]} for x in vals],
            })
    conflicts.sort(key=lambda x: -x["spread"])
    return cost, fee, conflicts, {}


# ---------------------------------------------------------------- 마진율표 형식
# "2026년 유통 마진율 및 예외단가리스트.xlsx" 처럼 시트마다 표가 흩어져 있는 파일을 읽는다.
# 규칙
#   원가        : [공급처옵션 | 상품코드] + [단가]        → 기재날짜가 가장 최근인 값 사용
#   이름 브릿지  : 코드가 있는 행에서 (상품명, 옵션명) → 코드 대응표를 만든다
#   코드 없는 표 : 위 브릿지로 코드를 찾아 원가/요율을 붙인다
#   셀러수수료   : [셀러 마진율]  (스룩 상품별 마진율)
#                 [할인 마진율]  (업체별 예외단가 — 업체명에 '스룩' 이 들어간 행만) ← 우선
MARGIN_MARKER_SHEETS = ("유통상품별마진율", "스룩_마진율1", "업체별 예외단가", "B2B상품별마진율")

_M = {
    "code": ["공급처옵션", "상품코드"],
    "cost": ["단가"],
    "name": ["상품명"],
    "opt": ["옵션명"],
    "date": ["기재날짜"],
    "seller_fee": ["셀러 마진율", "셀러마진율"],
    "disc_fee": ["할인 마진율", "할인마진율"],
    "vendor": ["업체", "업체명"],
}
_FAR_PAST = (0, 0, 0)


def month_cutoff(period):
    """'2026-07' → (2026, 7, 31) 형태의 상한. 그 달 말일까지만 유효한 단가로 본다."""
    try:
        y, m = int(str(period)[:4]), int(str(period)[5:7])
    except (ValueError, TypeError):
        return None
    return (y, m, 31)


def _as_date(v):
    """정렬 가능한 날짜 튜플. 날짜가 아니면 아주 과거로 취급."""
    try:
        return (v.year, v.month, v.day)
    except AttributeError:
        return _FAR_PAST


def locate_columns(rows, scan=40):
    """헤더가 여러 줄에 걸쳐 있어도 각 열을 찾아낸다.
       반환: (데이터 시작행 index, {필드: 열 index})"""
    found, hrow = {}, -1
    for r in range(min(scan, len(rows))):
        seen = {}
        for c, v in enumerate(rows[r]):
            k = norm(v)
            if k and k not in seen:
                seen[k] = c
        for field, aliases in _M.items():
            if field in found:
                continue
            for a in aliases:
                if norm(a) in seen:
                    found[field] = seen[norm(a)]
                    hrow = max(hrow, r)
                    break
    return hrow + 1, found


def _cell(row, col):
    return row[col] if col is not None and col < len(row) else None


def _date_txt(v):
    try:
        return "{:04d}-{:02d}-{:02d}".format(v.year, v.month, v.day)
    except AttributeError:
        return ""


def build_conflicts(occ, cost_best):
    """같은 상품코드가 서로 다른 원가로 여러 번 등록된 건을 정리한다.
       반환: [{code, used, used_date, values, rows:[...최신순...]}, ...]
       금액 차이가 큰 것부터 정렬."""
    out = []
    for k, recs in occ.items():
        vals = {r["value"] for r in recs}
        if len(vals) < 2:
            continue
        used_d, used_v = cost_best[k][0], cost_best[k][1]
        seen, uniq = set(), []          # 완전히 똑같은 행은 한 번만 보여준다
        for r in recs:
            sig = (r["date"], r["sheet"], r["name"], r["opt"], r["value"])
            if sig in seen:
                continue
            seen.add(sig)
            uniq.append(r)
        rows = sorted(uniq, key=lambda r: (r["date"], r["value"]), reverse=True)
        picked = False
        for r in rows:                       # 실제로 채택된 행 하나만 표시
            r["used"] = (not picked and r["date"] == used_d and r["value"] == used_v)
            picked = picked or r["used"]
        out.append({
            "code": k, "used": used_v, "used_date": _fmt_date_tuple(used_d),
            "values": sorted(vals), "spread": max(vals) - min(vals), "rows": rows,
        })
    out.sort(key=lambda x: -x["spread"])
    return out


def _fmt_date_tuple(d):
    return "" if d == _FAR_PAST else "{:04d}-{:02d}-{:02d}".format(*d)


def conflict_rows(conflicts):
    """중복 건을 화면·엑셀에 뿌리기 좋은 평평한 행 목록으로"""
    out = []
    for c in conflicts:
        for r in c["rows"]:
            out.append({
                "상품코드": c["code"],
                "사용": "★" if r.get("used") else "",
                "기재날짜": r["date_txt"],
                "시트": r["sheet"],
                "상품명": r["name"],
                "옵션명": r["opt"],
                "원가": r["value"],
                "최고-최저": c["spread"],
            })
    return out


def is_margin_workbook(wb):
    return any(n in wb.sheetnames for n in MARGIN_MARKER_SHEETS)


def load_margin_master(src, as_of=None):
    """마진율표 형식 → (cost, fee, conflicts, sheet_of_code)
       as_of: (y,m,d) 상한. 이보다 뒤에 기재된 줄은 무시한다
              (7월 이익률에 8월 단가를 쓰면 안 되므로)."""
    wb = open_wb(src, read_only=True, data_only=True)
    try:
        sheets = [(ws.title, list(ws.iter_rows(values_only=True))) for ws in wb.worksheets]
    finally:
        wb.close()

    parsed = []
    for title, rows in sheets:
        start, col = locate_columns(rows)
        parsed.append((title, rows, start, col))

    # 1) 코드가 있는 표 → 원가 + 이름 브릿지
    #    occ[코드] = 그 코드가 나온 모든 행 (시트·날짜·상품명·옵션명·단가) — 중복 확인용
    cost_best, occ, bridge = {}, defaultdict(list), {}
    nameless = []
    for title, rows, start, col in parsed:
        if "cost" not in col:
            continue
        has_code = "code" in col
        for i, row in enumerate(rows[start:], start=start + 1):
            raw = _cell(row, col["cost"])
            if raw is None or (isinstance(raw, str) and raw.strip().startswith("#")):
                continue
            v = num(raw)
            if v <= 0:
                continue
            d_raw = _cell(row, col.get("date"))
            d = _as_date(d_raw)
            if as_of and d != _FAR_PAST and d > as_of:
                continue                     # 기준월 이후 기재분은 제외
            nm, op = _cell(row, col.get("name")), _cell(row, col.get("opt"))
            pair = (norm(nm), norm(op)) if nm and op else None
            rec = {"sheet": title, "row": i, "date": d, "date_txt": _date_txt(d_raw),
                   "name": str(nm or "").strip(), "opt": str(op or "").strip(), "value": v}
            # (아래에서 cost_best 를 갱신할 때 시트명도 함께 남긴다)
            if has_code:
                k = code_key(_cell(row, col["code"]))
                if not k:
                    continue
                occ[k].append(rec)
                if k not in cost_best or d > cost_best[k][0]:
                    cost_best[k] = (d, v, title)
                if pair:
                    bridge.setdefault(pair, k)
            elif pair:
                nameless.append((pair, rec))

    # 2) 코드 없는 표 → 브릿지로 코드를 찾아 원가 보완
    for pair, rec in nameless:
        k = bridge.get(pair)
        if not k:
            continue
        occ[k].append(rec)
        if k not in cost_best or rec["date"] > cost_best[k][0]:
            cost_best[k] = (rec["date"], rec["value"], rec["sheet"])

    # 3) 셀러수수료 — 예외단가(할인 마진율)가 기본 마진율보다 우선
    base_fee, exc_fee = {}, {}
    for title, rows, start, col in parsed:
        for field, target, only_slook in (("seller_fee", base_fee, False),
                                          ("disc_fee", exc_fee, True)):
            if field not in col:
                continue
            for row in rows[start:]:
                r = rate(_cell(row, col[field]))
                if not r or r <= 0:
                    continue
                if only_slook and "스룩" not in str(_cell(row, col.get("vendor")) or ""):
                    continue
                k = code_key(_cell(row, col["code"])) if "code" in col else ""
                if not k:
                    nm, op = _cell(row, col.get("name")), _cell(row, col.get("opt"))
                    if not (nm and op):
                        continue
                    k = bridge.get((norm(nm), norm(op)), "")
                if k:
                    target[k] = r

    cost = {k: v for k, (_d, v, _t) in cost_best.items()}
    src = {k: t for k, (_d, _v, t) in cost_best.items()}
    fee = dict(base_fee)
    fee.update(exc_fee)
    conflicts = build_conflicts(occ, cost_best)

    # 4) 상품명+옵션명 → 코드 색인 (샘플비용용).  세트/낱개 표기를 걷어낸 열쇠도 함께 만든다
    #    (기준월 상한과 무관하게 만든다 — 코드를 찾는 색인일 뿐이고,
    #     실제 단가는 상한이 적용된 cost 에서 가져오므로 안전하다)
    names = {"exact": {}, "core": {}}
    for title, rows, start, col in parsed:
        if not ("code" in col and "name" in col and "opt" in col):
            continue
        for row in rows[start:]:
            k = code_key(_cell(row, col["code"]))
            nm, op = _cell(row, col["name"]), _cell(row, col["opt"])
            if not k or not nm or not op:
                continue
            names["exact"].setdefault((norm(nm), norm(op)), k)
            ck = (norm(nm), opt_core(op))
            if ck[1] and ck not in names["core"]:
                names["core"][ck] = (k, set_count(op) or set_count(nm) or 1)
    return cost, fee, conflicts, src, names


# ---------------------------------------------------------------- 연도별 파일 자동 인식
# 마진율표는 해마다 새 파일이 생기고, 지난해 파일은 그대로 보관된다.
#   2026년 유통 마진율 및 예외단가리스트.xlsx      ← 올해 (매월 갱신)
#   2025년\2025년 유통 마진율 및 예외단가리스트.xlsx  ← 작년 (고정)
# 폴더만 지정하면 파일 이름의 연도를 읽어 **최신 연도가 우선**하도록 자동 정렬한다.
YEAR_FILE_RE = re.compile(r"^\s*(20\d{2})\s*년\s*.*마진율")
DEFAULT_YEARS = 2          # 올해 + 작년


def find_year_files(folder, years=DEFAULT_YEARS):
    """폴더(및 바로 아래 하위폴더)에서 연도별 마진율표를 찾아 최신 연도 순으로 반환.
       반환: [(연도, Path), ...]"""
    from pathlib import Path as _P

    folder = _P(folder)
    if not folder.is_dir():
        return []
    cands = list(folder.glob("*.xls[xm]"))
    for sub in folder.iterdir():
        if sub.is_dir():
            cands += list(sub.glob("*.xls[xm]"))

    by_year = {}
    for f in cands:
        if f.name.startswith("~$"):
            continue
        m = YEAR_FILE_RE.match(f.name)
        if not m:
            continue
        y = int(m.group(1))
        # 같은 연도가 여러 개면 큰 파일(정식본) 우선
        prev = by_year.get(y)
        if prev is None or f.stat().st_size > prev.stat().st_size:
            by_year[y] = f
    out = sorted(by_year.items(), key=lambda kv: -kv[0])
    return out[:years] if years else out


def expand_sources(entries, base, years=DEFAULT_YEARS):
    """기준파일 목록에서 '폴더'로 적힌 항목을 연도별 파일로 펼친다.
       entries: [(이름, 경로문자열), ...]  →  [(라벨, Path), ...]  (없는 파일은 제외)
       두 번째 반환값은 화면에 보여줄 안내 메시지."""
    from pathlib import Path as _P

    out, notes = [], []
    for name, raw in entries:
        p = _P(str(raw).strip().strip('"'))
        if not p.is_absolute():
            p = _P(base) / p
        if p.is_dir():
            found = find_year_files(p, years)
            if not found:
                notes.append("{} — 폴더에서 연도별 마진율표를 찾지 못했습니다: {}".format(name, p))
            for y, f in found:
                out.append(("{}년".format(y), f))
            if found:
                notes.append("{} — 폴더에서 {} 인식 (앞쪽이 우선)".format(
                    name, " → ".join("{}년".format(y) for y, _f in found)))
        elif p.exists():
            out.append((name, p))
        else:
            notes.append("{} — 파일 없음: {}".format(name, p))
    return out, notes


def collect_candidates(sources):
    """모든 기준 파일에서 후보 단가를 전부 뽑는다 (DB 보관 → 월별 재계산용).
       반환: [(코드, 단가, 출처, 기재날짜 'YYYY-MM-DD' 또는 None, 파일순번, 줄번호)]"""
    from pathlib import Path as _P
    out = []
    for prio, (label, src) in enumerate(sources):
        wb = open_wb(src, read_only=True, data_only=True)
        try:
            if not is_margin_workbook(wb):
                continue
            sheets = [(ws.title, list(ws.iter_rows(values_only=True))) for ws in wb.worksheets]
        finally:
            wb.close()
        n = 0
        for title, rows in sheets:
            start, col = locate_columns(rows)
            if "cost" not in col or "code" not in col:
                continue
            for row in rows[start:]:
                k = code_key(_cell(row, col["code"]))
                raw = _cell(row, col["cost"])
                if not k or raw is None or (isinstance(raw, str) and str(raw).strip().startswith("#")):
                    continue
                v = num(raw)
                if v <= 0:
                    continue
                d = _date_txt(_cell(row, col.get("date"))) or None
                out.append((k, v, "{}·{}".format(label, title), d, prio, n))
                n += 1
    return out


def load_lookups(sources, as_of=None):
    """여러 기준 파일을 순서대로 읽어 합친다 (앞선 파일이 우선).
       sources: [(라벨, 경로 또는 파일객체), ...]
       반환: (cost, fee, conflicts, report, origin, names)
       origin: {상품코드: 그 값을 제공한 파일 라벨}
       names : 상품명+옵션명 → 코드 색인 (샘플비용 계산용)"""
    cost, fee, conflicts, report, origin = {}, {}, [], [], {}
    names = {"exact": {}, "core": {}}
    for label, src in sources:
        if src is None:
            continue
        wb = open_wb(src, read_only=True, data_only=True)
        margin = is_margin_workbook(wb)
        wb.close()
        if margin:
            c, f, cf, sheet_of, nx = load_margin_master(src, as_of)
        else:
            c, f, cf, sheet_of = load_master(src)
            nx = {"exact": {}, "core": {}}
        for _kind in ("exact", "core"):
            for _k, _v in nx[_kind].items():
                names[_kind].setdefault(_k, _v)
        added_c = sum(1 for k in c if k not in cost)
        added_f = sum(1 for k in f if k not in fee)
        for k, v in c.items():
            if k not in cost:
                cost[k] = v
                sh = sheet_of.get(k)
                origin[k] = "{}·{}".format(label, sh) if sh else label
        for k, v in f.items():
            fee.setdefault(k, v)
        conflicts += cf
        report.append({"label": label, "format": "마진율표" if margin else "구마스터",
                       "cost": len(c), "fee": len(f),
                       "cost_new": added_c, "fee_new": added_f})
    return cost, fee, conflicts, report, origin, names


# ---------------------------------------------------------------- 샘플비용
# 샘플 리스트의 옵션명은 **낱개** 기준이고, 마진율표는 **세트** 기준인 경우가 많다.
#   샘플 [7호-나시-단품]        ↔  마진율표 [7호-나시-3개1세트]   → 세트단가 ÷ 3
#   샘플 [110-화이트]-낱개 1개   ↔  마진율표 [110-화이트]           → 상품명이 '3개1세트' → ÷ 3
# 그래서 세트/낱개 표기를 걷어낸 '핵심 옵션명'으로 한 번 더 맞춰 본다.
SET_RE = re.compile(r"(\d+)\s*개\s*1?\s*세트")
LOOSE_RE = re.compile(r"낱개\s*(\d+)?\s*개?")
SOLO_RE = re.compile(r"단품")
_TRIM = " -·,/[]()"

SAMPLE_ALIASES = {
    "date": ["날짜", "일자", "주문일", "발주일", "지원일"],
    "gubun": ["구분", "부담구분", "비용구분"],
    "name": ["상품명", "품목명"],
    "opt": ["옵션명", "옵션"],
}
SAMPLE_OPTIONAL = {
    "qty": ["수량"],
    "ship": ["배송비"],
    "price": ["상품단가", "단가"],
    "state": ["진행여부"],
}
OWN_COST = "자체부담"          # 이 구분만 원가에 반영한다


def set_count(text):
    """'3개1세트' → 3 (없으면 0)"""
    m = SET_RE.search(norm(text))
    return int(m.group(1)) if m else 0


def loose_count(text):
    """'낱개 2개' → 2 · '낱개'/'단품' → 1 · 표기 없으면 0"""
    n = norm(text)
    m = LOOSE_RE.search(n)
    if m:
        return int(m.group(1)) if m.group(1) else 1
    return 1 if SOLO_RE.search(n) else 0


def opt_core(text):
    """옵션명에서 세트·낱개·단품 표기를 걷어낸 핵심 문자열"""
    s = norm(text)
    s = SET_RE.sub("", s)
    s = LOOSE_RE.sub("", s)
    s = SOLO_RE.sub("", s)
    prev = None
    while s != prev:
        prev = s
        s = s.strip(_TRIM)
    return s


def read_samples(src):
    """샘플 발주/지원 리스트(첫 번째 시트) → 줄 목록.
       반환: (rows, 시트이름).  rows 항목 = dict(month, date, gubun, name, opt, qty, ship, price, state)"""
    # 읽기전용 모드 — 이 파일은 시트 크기를 100만 행으로 적어두는 경우가 많아
    # 일반 모드로 열면 빈 셀 수천만 개를 훑느라 1분 가까이 걸린다.
    wb = open_wb(src, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        title = ws.title
        rows_all = list(ws.iter_rows(values_only=True))
        hr, col = find_header(ws, SAMPLE_ALIASES, rows=rows_all)
        if hr is None:
            seen = [str(v).strip() for v in (rows_all[0] if rows_all else [])
                    if v is not None and str(v).strip()]
            raise ValueError(
                "샘플 파일에서 열을 찾지 못했습니다. 첫 시트 '{}' 에 "
                "날짜·구분·상품명·옵션명 열이 있어야 합니다.  찾은 열: {}".format(
                    title, ", ".join(seen[:14]) or "(없음)"))
        _, opt = find_header(ws, SAMPLE_OPTIONAL, rows=rows_all)
        if opt:
            col.update(opt)
        out = []
        for row in rows_all[hr:]:
            nm = _pick(row, col, "name")
            if not nm or not str(nm).strip():
                continue
            d = _as_date(_pick(row, col, "date"))
            if d == _FAR_PAST:
                continue
            out.append({
                "month": "{:04d}-{:02d}".format(d[0], d[1]),
                "date": d,
                "gubun": str(_pick(row, col, "gubun") or "").strip(),
                "name": str(nm).strip(),
                "opt": str(_pick(row, col, "opt") or "").strip(),
                "qty": num(_pick(row, col, "qty")) or 1.0,
                "ship": num(_pick(row, col, "ship")),
                "price": num(_pick(row, col, "price")),
                "state": str(_pick(row, col, "state") or "").strip(),
            })
        return out, title
    finally:
        wb.close()


def lookup_sample_cost(name, opt, names, cost):
    """상품명+옵션명으로 기준 원가를 찾는다.
       반환: (낱개원가, 상품코드, 찾은방법)  — 못 찾으면 (0, '', '')"""
    if not names:
        return 0.0, "", ""
    nn = norm(name)
    k = names.get("exact", {}).get((nn, norm(opt)))
    if k and cost.get(k):
        return float(cost[k]), k, "정확"
    core = opt_core(opt)
    if core:
        hit = names.get("core", {}).get((nn, core))
        if hit:
            k, setn = hit
            v = cost.get(k)
            if v:
                if loose_count(opt) and setn > 1:
                    return float(v) / setn, k, "세트÷{}".format(setn)
                return float(v), k, "옵션정리"
    return 0.0, "", ""


def match_samples(rows, period, names, cost, own_only=True):
    """샘플 줄 → 기준월 집계.
       반환 dict:
         items    매칭된 줄 [{name, opt, qty, unit, amount, code, how}]
         missing  못 찾은 줄 [{name, opt, qty, price}]  (price=파일에 적힌 단가)
         goods    상품원가 합계
         ship     배송비 합계
         months   {월: {rows, goods, ship, miss}}  — 파일 전체 기준"""
    months = {}
    items, missing = [], []
    goods = ship = 0.0
    for r in rows:
        if own_only and norm(r["gubun"]) != OWN_COST:
            continue
        m = months.setdefault(r["month"], {"rows": 0, "goods": 0.0, "ship": 0.0, "miss": 0})
        unit, code, how = lookup_sample_cost(r["name"], r["opt"], names, cost)
        mult = loose_count(r["opt"]) or 1
        amount = unit * mult * (r["qty"] or 1)
        m["rows"] += 1
        m["ship"] += r["ship"]
        if unit:
            m["goods"] += amount
        else:
            m["miss"] += 1
        if r["month"] != period:
            continue
        ship += r["ship"]
        if unit:
            goods += amount
            items.append({"name": r["name"], "opt": r["opt"], "qty": r["qty"],
                          "unit": unit, "amount": amount, "code": code, "how": how})
        else:
            missing.append({"name": r["name"], "opt": r["opt"], "qty": r["qty"],
                            "price": r["price"]})
    return {"items": items, "missing": missing, "goods": goods, "ship": ship,
            "months": months}


def load_overrides(src):
    """원가보정 — 반환: (원가보정 {code: (값, 우선적용)}, 셀러수수료보정 {code: 율})"""
    if src is None:
        return {}, {}
    wb = open_wb(src, read_only=True, data_only=True)
    try:
        ov = {}
        ws = wb["원가보정"] if "원가보정" in wb.sheetnames else wb.worksheets[0]
        hr, col = find_header(ws, {"code": ["상품코드"], "cost": ["원가"]})
        if hr is not None:
            _, opt = find_header(ws, {"force": ["우선적용"]})
            if opt:
                col.update(opt)
            for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                k = code_key(_pick(row, col, "code"))
                raw = _pick(row, col, "cost")
                if k and raw is not None and str(raw).strip() != "":
                    force = is_yes(_pick(row, col, "force")) if "force" in col else True
                    ov[k] = (num(raw), force)

        fv = {}
        if "셀러수수료보정" in wb.sheetnames:
            ws = wb["셀러수수료보정"]
            hr, col = find_header(ws, {"code": ["상품코드"], "fee": ["셀러수수료"]})
            if hr is not None:
                for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                    k = code_key(_pick(row, col, "code"))
                    raw = _pick(row, col, "fee")
                    if k and raw is not None and str(raw).strip() != "":
                        r = rate(raw)
                        if r is not None:
                            fv[k] = r
    finally:
        wb.close()
    return ov, fv


# ---------------------------------------------------------------- 매출 파일
def _row(get):
    return {
        "group_code": str(get("group_code") or "").strip(),
        "group_name": str(get("group_name") or "").strip(),
        "code": code_key(get("code")),
        "code_raw": str(get("code") or "").strip(),
        "name": str(get("name") or "").strip(),
        "qty": num(get("qty")),
        "amount": num(get("amount")),
    }


def read_sales(src, filename=""):
    """이카운트 다운로드(xlsx/csv) → 행 리스트"""
    if str(filename or getattr(src, "name", src)).lower().endswith(".csv"):
        return _read_sales_csv(src, filename)

    wb = open_wb(src, read_only=True, data_only=True)
    best = None
    seen_sheets = []
    try:
        for ws in wb.worksheets:
            need = {k: COLUMN_ALIASES[k] for k in ("code", "qty", "amount")}
            hr, col = find_header(ws, need)
            seen_sheets.append((ws.title, hr, ws.max_row or "?", ws.max_column or "?",
                                _peek_headers(ws)))
            if hr is None:
                continue
            _, opt = find_header(ws, {k: COLUMN_ALIASES[k] for k in ("group_code", "group_name", "name")})
            if opt:
                col.update(opt)
            rows = []
            for raw in ws.iter_rows(min_row=hr + 1, values_only=True):
                get = lambda f: _pick(raw, col, f)
                if not code_key(get("code")) or is_junk(get("code"), get("name")):
                    continue
                rows.append(_row(get))
            if rows and (best is None or len(rows) > len(best)):
                best = rows
    finally:
        wb.close()
    if best is None:
        raise ValueError(_why_failed(seen_sheets))
    return best


def _peek_headers(ws, scan=25):
    """헤더로 보이는 줄(글자가 2개 이상인 행) 중 가장 그럴듯한 것"""
    best, score = "", 0
    for r, row in enumerate(ws.iter_rows(max_row=scan, values_only=True), start=1):
        txt = [str(v).strip() for v in row if isinstance(v, str) and str(v).strip()]
        if len(txt) > score:
            score, best = len(txt), " | ".join(txt[:12])
    return best


def _why_failed(seen_sheets):
    """왜 못 읽었는지 화면에 그대로 보여줄 메시지"""
    lines = ["[품목명[규격]코드 · 수량 · 판매액] 열을 찾지 못했습니다.",
             "파일에서 찾은 내용:"]
    for title, hr, nrow, ncol, peek in seen_sheets[:6]:
        if hr is not None:
            lines.append("  · '{}' — 헤더는 {}행에서 찾았지만 데이터 행이 없습니다".format(title, hr))
        else:
            lines.append("  · '{}' ({}행 x {}열) 의 열 이름: {}".format(
                title, nrow, ncol, peek if peek else "(글자 없음)"))
    if not seen_sheets:
        lines.append("  · 시트가 없습니다.")
    lines.append("위 열 이름이 실제와 다르면 알려주세요 — 인식하도록 추가하겠습니다.")
    return "\n".join(lines)


def _read_sales_csv(src, filename=""):
    import csv
    import io

    if hasattr(src, "read"):
        src.seek(0)
        data = src.read()
    else:
        with open(src, "rb") as f:
            data = f.read()
    table = None
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            table = list(csv.reader(io.StringIO(data.decode(enc), newline="")))
            break
        except UnicodeDecodeError:
            continue
    if table is None:
        raise ValueError("csv 인코딩을 읽을 수 없습니다 (utf-8 / cp949 지원).")

    for hr, line in enumerate(table[:40]):
        seen = {norm(v): i + 1 for i, v in enumerate(line)}
        col = {}
        for field, aliases in COLUMN_ALIASES.items():
            for a in aliases:
                if norm(a) in seen:
                    col[field] = seen[norm(a)]
                    break
        if all(f in col for f in ("code", "qty", "amount")):
            rows = []
            for line2 in table[hr + 1:]:
                get = lambda f: _pick(line2, col, f)
                if not code_key(get("code")) or is_junk(get("code"), get("name")):
                    continue
                rows.append(_row(get))
            return rows
    raise ValueError("csv 에서 헤더를 찾지 못했습니다.")


# ---------------------------------------------------------------- 계산
def calc_channel(ch, rows, cost, fee, override, fee_override, origin=None):
    goods, delivery, missing = [], [], []
    for r in rows:
        if is_delivery(r):
            delivery.append(r)
            continue
        c = dict(r)
        ovv, force = override.get(c["code"], (None, False))
        mst = cost.get(c["code"])
        c["cost_master"] = mst
        if ovv is not None and (force or mst is None):
            c["cost"], c["cost_src"] = ovv, "보정"
        elif mst is not None:
            c["cost"] = mst
            c["cost_src"] = (origin or {}).get(c["code"], "기준표")
        else:
            c["cost"], c["cost_src"] = 0.0, "미매칭"
            missing.append(c)
        c["cost_sum"] = c["cost"] * c["qty"]

        if ch["seller"]:
            if c["code"] in fee_override:
                c["fee_rate"], c["fee_src"] = fee_override[c["code"]], "보정"
            elif c["code"] in fee:
                c["fee_rate"], c["fee_src"] = fee[c["code"]], "마스터"
            else:
                c["fee_rate"], c["fee_src"] = ch.get("fee_base", 0.0), "기본율"
            c["fee_missing"] = c["fee_src"] == "기본율"
            c["fee_sum"] = c["fee_rate"] * c["amount"]
        else:
            c["fee_rate"], c["fee_sum"] = None, 0.0
            c["fee_src"], c["fee_missing"] = "", False
        goods.append(c)

    return {
        "ch": ch,
        "goods": goods,
        "delivery": delivery,
        "missing": missing,
        "sales": sum(r["amount"] for r in goods),
        "cost": sum(r["cost_sum"] for r in goods),
        "fee": sum(r["fee_sum"] for r in goods),
        "delivery_sales": sum(r["amount"] for r in delivery),
    }


def apply_totals(results, card_total, sample):
    """채널별 카드수수료·샘플비용 배분 후 이익액 확정. 합계 dict 반환."""
    card_left, sample_left = card_total, sample
    total = defaultdict(float)
    for res in results:
        ch = res["ch"]
        card = card_left if ch["card"] else 0.0
        card_left -= card
        smp = sample_left if ch["sample"] else 0.0
        sample_left -= smp
        res["card"], res["sample"] = card, smp
        res["profit"] = res["sales"] - res["cost"] - res["fee"] - card - smp
        total["sales"] += res["sales"]
        total["cost"] += res["cost"]
        total["fee"] += res["fee"]
        total["card"] += card
        total["sample"] += smp
        total["profit"] += res["profit"]
        total["delivery"] += res["delivery_sales"]
    total["gross"] = total["sales"] + total["delivery"]
    total["margin"] = total["profit"] / total["gross"] if total["gross"] else 0.0
    return total


def card_rows_total(cards):
    """[(결제수단, 금액, 요율)] → [(수단, 금액, 요율, 수수료)], 합계"""
    out, tot = [], 0.0
    for m, amt, rt in cards:
        f = num(amt) * (rate(rt) or 0.0)
        out.append((m, num(amt), rate(rt) or 0.0, f))
        tot += f
    return out, tot


# ---------------------------------------------------------------- 엑셀 리포트
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR = PatternFill("solid", fgColor="D9E1F2")
WARN = PatternFill("solid", fgColor="FFF2CC")
BAD = PatternFill("solid", fgColor="FCE4E4")
MONEY = "#,##0"
PCT = "0.00%"


def put(ws, r, c, v, bold=False, fill=None, fmt=None, box=True, align=None):
    cell = ws.cell(r, c, v)
    if bold:
        cell.font = Font(bold=True)
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if box:
        cell.border = BOX
    if align:
        cell.alignment = Alignment(horizontal=align)
    return cell


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _summary(wb, results, total, cards, card_total, sample_items, sample, period):
    ws = wb.active
    ws.title = "종합"
    ws.cell(1, 1, "{} 이익률".format(period)).font = Font(bold=True, size=14)

    hdr = ["구분", "상품매출", "상품원가", "셀러수수료", "카드수수료", "샘플비용",
           "이익액", "이익률(상품)", "배송매출"]
    for i, h in enumerate(hdr, start=1):
        put(ws, 3, i, h, bold=True, fill=HDR, align="center")

    r = 3
    for res in results:
        r += 1
        put(ws, r, 1, res["ch"]["name"], bold=True)
        put(ws, r, 2, res["sales"], fmt=MONEY)
        put(ws, r, 3, res["cost"], fmt=MONEY)
        put(ws, r, 4, res["fee"], fmt=MONEY)
        put(ws, r, 5, res["card"], fmt=MONEY)
        put(ws, r, 6, res["sample"], fmt=MONEY)
        put(ws, r, 7, res["profit"], fmt=MONEY, bold=True)
        put(ws, r, 8, res["profit"] / res["sales"] if res["sales"] else 0, fmt=PCT)
        put(ws, r, 9, res["delivery_sales"], fmt=MONEY)

    r += 1
    put(ws, r, 1, "합계", bold=True, fill=HDR)
    for i, k in enumerate(["sales", "cost", "fee", "card", "sample", "profit"], start=2):
        put(ws, r, i, total[k], bold=True, fill=HDR, fmt=MONEY)
    put(ws, r, 8, total["profit"] / total["sales"] if total["sales"] else 0,
        bold=True, fill=HDR, fmt=PCT)
    put(ws, r, 9, total["delivery"], bold=True, fill=HDR, fmt=MONEY)

    r += 2
    for lab, val, fmt in [("총매출(배송 포함)", total["gross"], MONEY),
                          ("총이익액", total["profit"], MONEY),
                          ("이익률(총매출 기준)", total["margin"], PCT)]:
        put(ws, r, 1, lab, bold=True, box=False)
        put(ws, r, 2, val, fmt=fmt, box=False)
        r += 1
    ws.cell(r, 1, "* 배송 이익 제외 (배송매출 = 배송원가로 처리)").font = Font(size=9, color="808080")

    r += 2
    put(ws, r, 1, "카드수수료 내역", bold=True, box=False)
    r += 1
    for i, h in enumerate(["결제수단", "정상금액", "수수료율", "수수료"], start=1):
        put(ws, r, i, h, bold=True, fill=HDR, align="center")
    for m, amt, rt, f in cards:
        r += 1
        put(ws, r, 1, m)
        put(ws, r, 2, amt, fmt=MONEY)
        put(ws, r, 3, rt, fmt=PCT)
        put(ws, r, 4, f, fmt=MONEY)
    r += 1
    put(ws, r, 1, "합계", bold=True, fill=HDR)
    put(ws, r, 2, sum(x[1] for x in cards), bold=True, fill=HDR, fmt=MONEY)
    put(ws, r, 3, "", fill=HDR)
    put(ws, r, 4, card_total, bold=True, fill=HDR, fmt=MONEY)

    r += 2
    put(ws, r, 1, "샘플비용 내역", bold=True, box=False)
    r += 1
    for i, h in enumerate(["항목", "금액"], start=1):
        put(ws, r, i, h, bold=True, fill=HDR, align="center")
    for it, a in sample_items:
        r += 1
        put(ws, r, 1, it)
        put(ws, r, 2, a, fmt=MONEY)
    r += 1
    put(ws, r, 1, "합계", bold=True, fill=HDR)
    put(ws, r, 2, sample, bold=True, fill=HDR, fmt=MONEY)

    autofit(ws, [22, 16, 16, 14, 14, 13, 16, 13, 14])
    ws.freeze_panes = "A4"


def _channel(wb, res):
    ch = res["ch"]
    ws = wb.create_sheet(ch["name"][:31])
    title = ch["name"] + "   이익 = 판매액 - 원가합"
    if ch["seller"]:
        title += " - 셀러수수료"
    if ch["card"]:
        title += " - 카드수수료"
    if ch["sample"]:
        title += " - 샘플비용"
    put(ws, 1, 1, title, bold=True, box=False)

    cols = ["품목그룹3코드", "품목그룹3", "상품코드", "품목명[규격]", "수량", "판매액",
            "원가", "원가합", "원가출처"]
    if ch["seller"]:
        cols += ["셀러수수료율", "요율출처", "셀러수수료"]
    cols += ["이익액", "이익률"]
    for i, h in enumerate(cols, start=1):
        put(ws, 3, i, h, bold=True, fill=HDR, align="center")

    r = 3
    for g in res["goods"]:
        r += 1
        put(ws, r, 1, g["group_code"])
        put(ws, r, 2, g["group_name"])
        put(ws, r, 3, g["code_raw"])
        put(ws, r, 4, g["name"])
        put(ws, r, 5, g["qty"], fmt=MONEY)
        put(ws, r, 6, g["amount"], fmt=MONEY)
        put(ws, r, 7, g["cost"], fmt=MONEY)
        put(ws, r, 8, g["cost_sum"], fmt=MONEY)
        sc = put(ws, r, 9, g["cost_src"], align="center")
        i = 10
        if ch["seller"]:
            put(ws, r, i, g["fee_rate"], fmt=PCT)
            sf = put(ws, r, i + 1, g["fee_src"], align="center")
            if g["fee_missing"]:
                ws.cell(r, i).fill = WARN
                sf.fill = WARN
            put(ws, r, i + 2, g["fee_sum"], fmt=MONEY)
            i += 3
        profit = g["amount"] - g["cost_sum"] - g["fee_sum"]
        put(ws, r, i, profit, fmt=MONEY)
        put(ws, r, i + 1, profit / g["amount"] if g["amount"] else 0, fmt=PCT)
        if g["cost_src"] == "미매칭":
            for j in range(1, len(cols) + 1):
                ws.cell(r, j).fill = BAD
        elif g["cost_src"] == "보정":
            sc.fill = WARN

    r += 1
    put(ws, r, 4, "합계", bold=True, fill=HDR, align="center")
    put(ws, r, 6, res["sales"], bold=True, fill=HDR, fmt=MONEY)
    put(ws, r, 8, res["cost"], bold=True, fill=HDR, fmt=MONEY)
    if ch["seller"]:
        put(ws, r, 12, res["fee"], bold=True, fill=HDR, fmt=MONEY)
    put(ws, r, len(cols) - 1, res["sales"] - res["cost"] - res["fee"],
        bold=True, fill=HDR, fmt=MONEY)

    if res["delivery"]:
        r += 2
        put(ws, r, 1, "배송비 (이익 계산 제외 — 매출 = 원가)", bold=True, box=False)
        r += 1
        for i, h in enumerate(["품목그룹3코드", "품목그룹3", "상품코드", "품목명[규격]", "수량", "판매액"], start=1):
            put(ws, r, i, h, bold=True, fill=HDR, align="center")
        for d in res["delivery"]:
            r += 1
            for i, k in enumerate(["group_code", "group_name", "code_raw", "name"], start=1):
                put(ws, r, i, d[k])
            put(ws, r, 5, d["qty"], fmt=MONEY)
            put(ws, r, 6, d["amount"], fmt=MONEY)
        r += 1
        put(ws, r, 4, "합계", bold=True, fill=HDR, align="center")
        put(ws, r, 6, res["delivery_sales"], bold=True, fill=HDR, fmt=MONEY)

    autofit(ws, [14, 26, 22, 46, 8, 13, 11, 13, 10, 12, 10, 13, 13, 10])
    ws.freeze_panes = "A4"


def missing_rows(results):
    """미매칭 집계 → [(코드, 표시코드, 품목명, 수량, 판매액, 채널들)] 판매액 큰 순"""
    agg, chans = {}, defaultdict(set)
    for res in results:
        for m in res["missing"]:
            a = agg.setdefault(m["code"], {"qty": 0.0, "amt": 0.0, "name": "", "raw": ""})
            a["qty"] += m["qty"]
            a["amt"] += m["amount"]
            a["name"] = m["name"] or a["name"]
            a["raw"] = m["code_raw"] or a["raw"]
            chans[m["code"]].add(res["ch"]["name"])
    return [(k, a["raw"], a["name"], a["qty"], a["amt"], ", ".join(sorted(chans[k])))
            for k, a in sorted(agg.items(), key=lambda x: -x[1]["amt"])]


def fee_default_rows(results):
    out = []
    for res in results:
        if not res["ch"]["seller"]:
            continue
        for g in res["goods"]:
            if g["fee_missing"]:
                out.append((res["ch"]["name"], g["code_raw"], g["name"],
                            g["amount"], g["fee_rate"], g["fee_sum"]))
    return sorted(out, key=lambda x: -x[3])


def applied_rows(results):
    """보정이 적용된 건. 마스터 값을 덮어쓴 건은 diff 가 None 이 아님."""
    out = []
    for res in results:
        for g in res["goods"]:
            if g["cost_src"] != "보정":
                continue
            d = None if g["cost_master"] is None else (g["cost"] - g["cost_master"]) * g["qty"]
            out.append((res["ch"]["name"], g["code_raw"], g["name"], g["qty"],
                        g["cost_master"], g["cost"], d))
    return out


def _missing(wb, results):
    ws = wb.create_sheet("미매칭")
    ws.cell(1, 1, "원가를 찾지 못한 상품 — 기준 파일에 추가하거나 원가보정에 입력하세요"
            ).font = Font(bold=True, color="C00000")
    for i, h in enumerate(["채널", "상품코드", "품목명[규격]", "수량", "판매액", "원가(직접입력)"], start=1):
        put(ws, 3, i, h, bold=True, fill=HDR, align="center")
    rows = missing_rows(results)
    r = 3
    for _k, raw, nm, qty, amt, chs in rows:
        r += 1
        put(ws, r, 1, chs)
        put(ws, r, 2, raw)
        put(ws, r, 3, nm)
        put(ws, r, 4, qty, fmt=MONEY)
        put(ws, r, 5, amt, fmt=MONEY)
        put(ws, r, 6, None, fill=WARN)
    if rows:
        r += 1
        put(ws, r, 3, "합계", bold=True, fill=HDR, align="center")
        put(ws, r, 5, sum(x[4] for x in rows), bold=True, fill=HDR, fmt=MONEY)
    else:
        put(ws, 4, 1, "없음 — 모든 상품의 원가를 찾았습니다.", box=False)
    autofit(ws, [16, 24, 50, 10, 14, 14])
    ws.freeze_panes = "A4"


def _fee_default(wb, results):
    ws = wb.create_sheet("셀러수수료_기본율적용")
    ws.cell(1, 1, "기준 파일에 셀러수수료율이 없어 [기본율]로 계산된 상품"
            ).font = Font(bold=True, color="C00000")
    for i, h in enumerate(["채널", "상품코드", "품목명[규격]", "판매액", "적용요율", "수수료"], start=1):
        put(ws, 3, i, h, bold=True, fill=HDR, align="center")
    rows = fee_default_rows(results)
    r = 3
    for ch, raw, nm, amt, rt, f in rows:
        r += 1
        put(ws, r, 1, ch)
        put(ws, r, 2, raw)
        put(ws, r, 3, nm)
        put(ws, r, 4, amt, fmt=MONEY)
        put(ws, r, 5, rt, fmt=PCT)
        put(ws, r, 6, f, fmt=MONEY)
    if rows:
        r += 1
        put(ws, r, 3, "합계", bold=True, fill=HDR, align="center")
        put(ws, r, 6, sum(x[5] for x in rows), bold=True, fill=HDR, fmt=MONEY)
    else:
        put(ws, 4, 1, "없음 — 모든 상품이 기준 파일에서 매칭되었습니다.", box=False)
    autofit(ws, [16, 24, 50, 14, 11, 14])
    ws.freeze_panes = "A4"


def _applied(wb, results):
    ws = wb.create_sheet("보정적용내역")
    ws.cell(1, 1, "원가보정이 실제로 적용된 상품").font = Font(bold=True)
    ws.cell(2, 1, "노란 줄 = 기준 파일에 값이 있는데도 덮어쓴 건. 어느 쪽이 맞는지 확인하세요."
            ).font = Font(size=9, color="C00000")
    for i, h in enumerate(["채널", "상품코드", "품목명[규격]", "수량",
                           "기준 원가", "적용 원가", "차이(원가합)"], start=1):
        put(ws, 3, i, h, bold=True, fill=HDR, align="center")
    rows = applied_rows(results)
    r, diff = 3, 0.0
    for ch, raw, nm, qty, mst, cst, d in rows:
        r += 1
        put(ws, r, 1, ch)
        put(ws, r, 2, raw)
        put(ws, r, 3, nm)
        put(ws, r, 4, qty, fmt=MONEY)
        put(ws, r, 5, mst, fmt=MONEY)
        put(ws, r, 6, cst, fmt=MONEY)
        put(ws, r, 7, d, fmt=MONEY)
        if d is not None:
            diff += d
            for j in range(1, 8):
                ws.cell(r, j).fill = WARN
    if rows:
        r += 1
        put(ws, r, 3, "마스터 대비 원가 증감", bold=True, fill=HDR, align="center")
        put(ws, r, 7, diff, bold=True, fill=HDR, fmt=MONEY)
    else:
        put(ws, 4, 1, "없음", box=False)
    autofit(ws, [14, 24, 46, 9, 13, 13, 15])
    ws.freeze_panes = "A4"


def _conflicts(wb, conflicts):
    ws = wb.create_sheet("중복코드점검")
    ws.cell(1, 1, "같은 상품코드가 서로 다른 원가로 여러 번 등록된 건"
            ).font = Font(bold=True, color="C00000")
    ws.cell(2, 1, "★ 표시가 실제로 사용된 값입니다 (기재날짜가 가장 최근인 것). "
                  "금액 차이가 큰 순서로 정렬했습니다."
            ).font = Font(size=9, color="808080")
    cols = ["상품코드", "사용", "기재날짜", "시트", "상품명", "옵션명", "원가", "최고-최저"]
    for i, h in enumerate(cols, start=1):
        put(ws, 4, i, h, bold=True, fill=HDR, align="center")

    r, prev = 4, None
    for row in conflict_rows(conflicts):
        r += 1
        newblock = row["상품코드"] != prev
        prev = row["상품코드"]
        put(ws, r, 1, row["상품코드"] if newblock else "")
        put(ws, r, 2, row["사용"], align="center")
        put(ws, r, 3, row["기재날짜"], align="center")
        put(ws, r, 4, row["시트"])
        put(ws, r, 5, row["상품명"])
        put(ws, r, 6, row["옵션명"])
        put(ws, r, 7, row["원가"], fmt=MONEY)
        put(ws, r, 8, row["최고-최저"] if newblock else None, fmt=MONEY)
        if row["사용"]:
            for j in range(1, len(cols) + 1):
                ws.cell(r, j).fill = WARN
        if newblock:
            top = Side(style="medium", color="9E9E9E")
            for j in range(1, len(cols) + 1):
                ws.cell(r, j).border = Border(left=THIN, right=THIN, bottom=THIN, top=top)
    if not conflicts:
        put(ws, 5, 1, "없음 — 중복 등록된 상품코드가 없습니다.", box=False)
    autofit(ws, [22, 6, 12, 20, 34, 30, 12, 12])
    ws.freeze_panes = "A5"


def build_report(results, total, cards, card_total, sample_items, sample, conflicts, period):
    """결과 openpyxl Workbook 생성 (호출측이 저장)"""
    wb = openpyxl.Workbook()
    _summary(wb, results, total, cards, card_total, sample_items, sample, period)
    for res in results:
        _channel(wb, res)
    _missing(wb, results)
    _fee_default(wb, results)
    _applied(wb, results)
    _conflicts(wb, conflicts)
    return wb
