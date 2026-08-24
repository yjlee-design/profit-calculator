# -*- coding: utf-8 -*-
"""
초기설정 — 입력/설정.xlsx 와 원가보정.xlsx 를 만든다.
이미 있으면 덮어쓰지 않는다. (--force 를 붙이면 새로 만듦)

원가보정.xlsx 는 `7월 이익률.xlsx` 에 수식 없이 손으로 입력된 원가를 자동으로 긁어와 채운다.
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
INPUT_DIR = BASE / "입력"
CONFIG = INPUT_DIR / "설정.xlsx"
OVERRIDE = BASE / "원가보정.xlsx"
SEED = BASE / "7월 이익률.xlsx"

FORCE = "--force" in sys.argv

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR = PatternFill("solid", fgColor="D9E1F2")
IN = PatternFill("solid", fgColor="FFF2CC")


def head(ws, row, names, widths=None):
    for i, n in enumerate(names, start=1):
        c = ws.cell(row, i, n)
        c.font = Font(bold=True)
        c.fill = HDR
        c.border = BOX
        c.alignment = Alignment(horizontal="center")
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def cell(ws, r, c, v, fmt=None, fill=None):
    x = ws.cell(r, c, v)
    x.border = BOX
    if fmt:
        x.number_format = fmt
    if fill:
        x.fill = fill
    return x


def make_config():
    INPUT_DIR.mkdir(exist_ok=True)
    if CONFIG.exists() and not FORCE:
        print("  설정.xlsx 이미 있음 — 건너뜀")
        return
    wb = openpyxl.Workbook()

    # 기간
    ws = wb.active
    ws.title = "기간"
    ws.cell(1, 1, "매달 이 값만 바꾸면 됩니다").font = Font(bold=True, size=12)
    head(ws, 3, ["항목", "값"], [16, 20])
    cell(ws, 4, 1, "기준월")
    cell(ws, 4, 2, "2026-08", fill=IN)
    c = ws.cell(6, 1, "* 기준월은 결과 파일 이름이 됩니다 (예: 결과/2026-08 이익률.xlsx)")
    c.font = Font(size=9, color="808080")

    # 기준파일 (원가·셀러수수료를 가져올 파일들 — 위에 있는 것이 우선)
    ws = wb.create_sheet("기준파일")
    ws.cell(1, 1, "원가·셀러수수료를 읽어올 파일. 위에 있는 파일이 우선합니다.").font = Font(bold=True)
    ws.cell(2, 1, "* 경로는 이 폴더 기준 파일명이거나, 전체 경로를 적어도 됩니다.").font = Font(size=9, color="808080")
    ws.cell(3, 1, r"  예) 공유폴더  \\서버이름\공유\2026년 유통 마진율 및 예외단가리스트.xlsx").font = Font(size=9, color="808080")
    ws.cell(4, 1, r"  예) OneDrive  C:\Users\사용자\OneDrive\마진율\2026년 유통 마진율.xlsx").font = Font(size=9, color="808080")
    ws.cell(5, 1, "* 실행할 때마다 그 경로의 최신 파일을 그대로 읽습니다 (복사 안 해도 됨).").font = Font(size=9, color="808080")
    head(ws, 7, ["순서", "이름", "파일경로", "사용"], [6, 16, 68, 8])
    for i, (nm, path) in enumerate([
        ("마진율표", "2026년 유통 마진율 및 예외단가리스트.xlsx"),
        ("구마스터", "이익률 마스터.xlsx"),
    ], start=8):
        cell(ws, i, 1, i - 7)
        cell(ws, i, 2, nm)
        cell(ws, i, 3, path, None, IN)
        c = cell(ws, i, 4, "O", None, IN)
        c.alignment = Alignment(horizontal="center")

    # 채널설정
    ws = wb.create_sheet("채널설정")
    ws.cell(1, 1, "이카운트에서 채널별로 받은 파일을 입력/ 폴더에 넣고, 파일명을 여기에 적으세요").font = Font(bold=True)
    head(ws, 3, ["채널명", "파일명", "셀러수수료적용", "카드수수료적용", "샘플비용적용",
                 "셀러수수료기본율"], [16, 22, 15, 15, 14, 16])
    rows = [
        ("유통", "유통.xlsx", "X", "X", "O", None),
        ("유통B2B", "유통B2B.xlsx", "X", "X", "X", None),
        ("유통_입점몰", "유통_입점몰.xlsx", "X", "X", "X", None),
        ("스룩", "스룩.xlsx", "O", "O", "X", 0.15),
    ]
    for i, row in enumerate(rows, start=4):
        for j, v in enumerate(row, start=1):
            c = cell(ws, i, j, v, "0.00%" if j == 6 else None, IN if j >= 3 else None)
            if j >= 3:
                c.alignment = Alignment(horizontal="center")
    notes = [
        "* O = 차감함 / X = 차감 안 함.  채널을 추가하려면 아래에 줄만 추가하세요.",
        "* 셀러수수료·카드수수료·샘플비용은 'O' 인 채널의 이익에서 차감됩니다.",
        "* 셀러수수료기본율: 마스터에서 요율을 못 찾은 상품에 적용할 기본값 (스룩 15%).",
    ]
    for k, t in enumerate(notes, start=9):
        ws.cell(k, 1, t).font = Font(size=9, color="808080")

    # 카드수수료 — 월별 누적
    ws = wb.create_sheet("카드수수료")
    ws.cell(1, 1, "스룩 결제수단별 금액 (월별로 쌓입니다)").font = Font(bold=True)
    ws.cell(2, 1, "* 웹앱에서 입력하면 자동으로 여기에 저장됩니다. 직접 적어도 됩니다.").font = Font(size=9, color="808080")
    head(ws, 4, ["기준월", "결제수단", "정상금액", "수수료율"], [12, 16, 16, 12])
    i = 5
    for m, rt in CARD_SEED:
        cell(ws, i, 1, "2026-07")
        cell(ws, i, 2, m)
        cell(ws, i, 3, CARD_JULY.get(m, 0), "#,##0", IN)
        cell(ws, i, 4, rt, "0.00%", IN)
        i += 1
    ws.freeze_panes = "A5"

    # 샘플비용 — 월별 누적
    ws = wb.create_sheet("샘플비용")
    ws.cell(1, 1, "자체 부담 비용 (월별로 쌓입니다)").font = Font(bold=True)
    ws.cell(2, 1, "* 웹앱에서 입력하면 자동으로 여기에 저장됩니다. 지난 달 값은 지우지 마세요.").font = Font(size=9, color="808080")
    head(ws, 4, ["기준월", "항목", "금액"], [12, 26, 16])
    for i, (mon, amt) in enumerate(SAMPLE_SEED, start=5):
        cell(ws, i, 1, mon)
        cell(ws, i, 2, "자체샘플+이벤트지원")
        cell(ws, i, 3, amt, "#,##0", IN)
    ws.freeze_panes = "A5"

    wb.save(CONFIG)
    print("  생성: 입력\\설정.xlsx")


RANGES = {"유통": (14, 1175), "스룩": (16, 563), "유통B2B": (13, 53), "유통_입점몰": (13, 33)}

# 결제수단과 기본 수수료율
CARD_SEED = [("신용카드", 0.0275), ("무통장입금", 0.0), ("페이코", 0.0275), ("토스", 0.0275)]
# 7월 이익률.xlsx 에서 옮겨온 값
CARD_JULY = {"신용카드": 79866800, "무통장입금": 4211100, "페이코": 3526200, "토스": 6827600}
SAMPLE_SEED = [("2026-01", 210348), ("2026-02", 115509), ("2026-03", 1120312),
               ("2026-04", 625220), ("2026-05", 1206017), ("2026-06", 743654),
               ("2026-07", 165304)]


def master_cost():
    """현재 마스터의 원가 조회표 (첫 값 우선)"""
    p = BASE / "이익률 마스터.xlsx"
    if not p.exists():
        return {}
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb["원가리스트_최종"]
    d = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 3 or row[1] is None:
            continue
        k = str(row[1]).strip().upper()
        if k and k != "상품코드" and k not in d:
            d[k] = row[2]
    wb.close()
    return d


def collect_seed(col_idx, skip_same):
    """7월 파일에서 수식 없이 직접 입력된 값을 추출.
       skip_same: 마스터와 값이 같으면 제외 (의미 없는 중복 방지)"""
    if not SEED.exists():
        return []
    wbf = openpyxl.load_workbook(SEED, data_only=False)
    wbv = openpyxl.load_workbook(SEED, data_only=True)
    seen, out = set(), []
    for name, (r0, r1) in RANGES.items():
        if name not in wbf.sheetnames:
            continue
        wf, wv = wbf[name], wbv[name]
        for r in range(r0, r1 + 1):
            f = wf.cell(r, col_idx).value
            if isinstance(f, str) and f.startswith("="):
                continue
            code = wv.cell(r, 3).value
            val = wv.cell(r, col_idx).value
            if not code or val in (None, 0, ""):
                continue
            k = str(code).strip().upper()
            if k in seen:
                continue
            seen.add(k)
            if skip_same is not None and k in skip_same:
                try:
                    if abs(float(skip_same[k]) - float(val)) < 1e-6:
                        continue                      # 마스터와 동일 → 불필요
                    note = "마스터와 다름(마스터 {:,})".format(skip_same[k])
                except (TypeError, ValueError):
                    note = "확인 필요"
            else:
                note = "마스터에 없음"
            out.append((str(code).strip(), val, str(wv.cell(r, 4).value or ""), name, note))
    wbf.close()
    wbv.close()
    return out


def make_override():
    if OVERRIDE.exists() and not FORCE:
        print("  원가보정.xlsx 이미 있음 — 건너뜀")
        return
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "원가보정"
    ws.cell(1, 1, "마스터에 없거나 값이 틀린 원가를 여기에 적습니다 (한 번 적으면 매달 자동 적용)").font = Font(bold=True)
    ws.cell(2, 1, "* 우선적용 O = 마스터에 값이 있어도 이 값을 씀 / X(빈칸) = 마스터에 없을 때만 씀").font = Font(size=9, color="808080")
    ws.cell(3, 1, "* 마스터가 갱신되어 값이 맞게 되면 해당 줄을 지우세요.").font = Font(size=9, color="808080")
    head(ws, 5, ["상품코드", "원가", "우선적용", "품목명(참고)", "채널", "비고"],
         [24, 12, 10, 50, 12, 30])
    seed = collect_seed(8, master_cost())
    for i, (code, val, nm, src, note) in enumerate(seed, start=6):
        cell(ws, i, 1, code)
        cell(ws, i, 2, val, "#,##0", IN)
        c = cell(ws, i, 3, "O" if note.startswith("마스터와 다름") else "X", None, IN)
        c.alignment = Alignment(horizontal="center")
        cell(ws, i, 4, nm)
        cell(ws, i, 5, src)
        cell(ws, i, 6, note)
    ws.freeze_panes = "A6"

    ws = wb.create_sheet("셀러수수료보정")
    ws.cell(1, 1, "마스터보다 우선 적용되는 셀러수수료율").font = Font(bold=True)
    ws.cell(2, 1, "* 기본율(설정.xlsx)과 다른 상품만 적어주세요. 15% → 0.15 또는 15% 로 입력.").font = Font(size=9, color="808080")
    head(ws, 4, ["상품코드", "셀러수수료", "품목명(참고)", "채널", "비고"], [24, 12, 50, 12, 26])
    fseed = [x for x in collect_seed(11, None) if abs(float(x[1]) - 0.15) > 1e-9]
    for i, (code, val, nm, src, _n) in enumerate(fseed, start=5):
        cell(ws, i, 1, code)
        cell(ws, i, 2, val, "0.00%", IN)
        cell(ws, i, 3, nm)
        cell(ws, i, 4, src)
        cell(ws, i, 5, "7월 수동입력 (기본율과 다름)")
    ws.freeze_panes = "A5"

    wb.save(OVERRIDE)
    print("  생성: 원가보정.xlsx  (원가 {}건 / 셀러수수료 {}건 이관)".format(len(seed), len(fseed)))


if __name__ == "__main__":
    print("초기설정 시작")
    make_config()
    make_override()
    (BASE / "결과").mkdir(exist_ok=True)
    print("완료.\n")
    print("다음 순서:")
    print("  1) 입력/설정.xlsx 에 기준월·카드수수료·샘플비용 입력")
    print("  2) 이카운트 매출 파일 4개를 입력/ 폴더에 넣기")
    print("  3) 실행.bat 더블클릭")
